# -*- coding: utf-8 -*-
import base64
import hashlib
import logging

import logging.handlers

import os
import re
import unicodedata

import subprocess

import sys

import tempfile
import ctypes
import threading
import secrets



import tkinter as tk

from tkinter import ttk, messagebox, scrolledtext, filedialog


def _show_missing_dependency_error(module_name):
    app_root = os.path.abspath(os.path.dirname(__file__))
    requirements_path = os.path.join(app_root, "requirements.txt")
    venv_python = os.path.join(app_root, ".venv", "Scripts", "python.exe")

    if os.path.exists(venv_python):
        install_python = venv_python
        run_python = venv_python
    else:
        install_python = sys.executable
        run_python = sys.executable

    install_cmd = f'"{install_python}" -m pip install -r "{requirements_path}"'
    run_cmd = f'"{run_python}" "{os.path.join(app_root, "app.py")}"'
    message = (
        f"Falta el modulo '{module_name}'.\n\n"
        "Instala las dependencias con:\n"
        f"{install_cmd}\n\n"
        "Luego ejecuta la app con:\n"
        f"{run_cmd}"
    )

    print(message, file=sys.stderr)
    if os.name == "nt":
        try:
            ctypes.windll.user32.MessageBoxW(0, message, "RECA Empresas", 0x10)
        except Exception:
            pass


try:
    import requests
    from dotenv import load_dotenv
    from openpyxl import load_workbook
    from supabase import create_client, Client
except ModuleNotFoundError as exc:
    _show_missing_dependency_error(exc.name)
    raise SystemExit(1) from exc



# ============================================

# CONFIGURACION

# ============================================



APP_NAME = "RECA Empresas"

APP_VERSION = "1.0.19"

GITHUB_OWNER = "auyaban"

GITHUB_REPO = "reca-empresas"

UPDATE_ASSET_NAME = "RECA_Setup.exe"
UPDATE_HASH_NAME = "RECA_Setup.exe.sha256"

COLOR_PURPLE = "#7C3D96"
COLOR_PURPLE_DARK = "#5E2D73"
COLOR_TEAL = "#07B499"
COLOR_TEAL_DARK = "#059680"
COLOR_LIGHT_BG = "#F7F5FA"

# --- Colores semanticos ---
COLOR_DANGER = "#dc3545"
COLOR_DANGER_DARK = "#c82333"
COLOR_SUCCESS = "#28a745"
COLOR_SUCCESS_DARK = "#218838"
COLOR_WARNING = "#FF9800"
COLOR_WARNING_DARK = "#E68900"
COLOR_INFO = "#1E88E5"
COLOR_INFO_DARK = "#1565C0"
COLOR_NEUTRAL = "#455A64"
COLOR_NEUTRAL_DARK = "#37474F"
COLOR_BORDER = "#E0DAE8"
COLOR_TEXT = "#2D2D3F"
COLOR_WHITE = "#FFFFFF"
TREE_ROW_ALT = "#F3F0F7"

# --- Tipografia ---
FONT_FAMILY = "Segoe UI" if sys.platform == "win32" else "Arial"
FONT_H1 = (FONT_FAMILY, 20, "bold")
FONT_H2 = (FONT_FAMILY, 16, "bold")
FONT_H3 = (FONT_FAMILY, 12, "bold")
FONT_BODY = (FONT_FAMILY, 11)
FONT_BODY_BOLD = (FONT_FAMILY, 11, "bold")
FONT_SMALL = (FONT_FAMILY, 9)

# --- Espaciado ---
SP_XS = 4
SP_SM = 8
SP_MD = 16
SP_LG = 24

# --- Botones ---
BTN_PADX = 16
BTN_PADY = 8

REQUEST_HEADERS = {"User-Agent": f"{APP_NAME}/{APP_VERSION}"}

DEFAULT_SUPABASE_AUTH_EMAIL = "test@reca.local"
DEFAULT_SUPABASE_AUTH_PASSWORD = "Reca.Test.2026!v3"
DEFAULT_PROFESIONAL_TEMP_PASSWORD = "Password1234"



def _resource_path(relative_path):
    base_dir = getattr(sys, "_MEIPASS", os.path.abspath(os.path.dirname(__file__)))
    return os.path.join(base_dir, relative_path)

LOGO_PATH = _resource_path(os.path.join("logo", "logo_reca.png"))


def _maximize_window(window):
    try:
        window.state("zoomed")
    except tk.TclError:
        try:
            window.attributes("-zoomed", True)
        except tk.TclError:
            pass


class ScrollableFrame(tk.Frame):
    """Contenedor scrollable reutilizable para formularios construidos con grid."""

    def __init__(self, parent, bg=None, **kwargs):
        super().__init__(parent, bg=bg, **kwargs)
        self._bg = bg
        canvas_bg = bg if bg is not None else self.cget("bg")
        self.canvas = tk.Canvas(self, bg=canvas_bg, highlightthickness=0, borderwidth=0)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.content = tk.Frame(self.canvas, bg=bg)

        self._content_window = self.canvas.create_window((0, 0), window=self.content, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        self.content.bind("<Configure>", self._on_content_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)

    def _on_content_configure(self, _event=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self._bind_mousewheel_recursive(self.content)

    def _on_canvas_configure(self, event):
        self.canvas.itemconfigure(self._content_window, width=event.width)

    def _bind_mousewheel_recursive(self, widget):
        if not getattr(widget, "_scrollable_mousewheel_bound", False):
            for sequence in ("<MouseWheel>", "<Button-4>", "<Button-5>"):
                widget.bind(sequence, self._on_mousewheel, add="+")
            widget._scrollable_mousewheel_bound = True
        for child in widget.winfo_children():
            self._bind_mousewheel_recursive(child)

    def _on_mousewheel(self, event):
        if event.num == 4:
            delta = -1
        elif event.num == 5:
            delta = 1
        else:
            delta = -1 * int(event.delta / 120) if event.delta else 0

        if delta:
            self.canvas.yview_scroll(delta, "units")
        return "break"


class SplashScreen(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("RECA")
        self.resizable(False, False)
        self.configure(bg="white")
        self.protocol("WM_DELETE_WINDOW", lambda: None)

        self.logo_image = None
        self._closed = False

        container = tk.Frame(self, bg="white")
        container.pack(fill=tk.BOTH, expand=True, padx=24, pady=20)

        if os.path.exists(LOGO_PATH):
            try:
                logo = tk.PhotoImage(file=LOGO_PATH)
                max_w, max_h = 240, 240
                scale = max(1, logo.width() // max_w, logo.height() // max_h)
                if scale > 1:
                    logo = logo.subsample(scale, scale)
                self.logo_image = logo
                tk.Label(container, image=self.logo_image, bg="white").pack(pady=(0, 12))
            except Exception:
                self.logo_image = None

        tk.Label(
            container,
            text="Bienvenido al sistema de gestion de empresas de RECA",
            font=FONT_H3,
            bg="white",
            fg=COLOR_PURPLE,
            wraplength=420,
            justify="center",
        ).pack(pady=(0, 12))

        self.status_label = tk.Label(
            container,
            text="Iniciando...",
            font=FONT_BODY,
            bg="white",
            fg=COLOR_TEAL,
        )
        self.status_label.pack(pady=(0, 8))

        self.progress = ttk.Progressbar(
            container,
            length=360,
            mode="determinate",
            maximum=100,
            style="Reca.Horizontal.TProgressbar",
        )
        self.progress.pack(pady=(0, 12))

        self.log_box = tk.Text(container, height=6, width=52, bg="#F2EFF6", bd=0)
        self.log_box.configure(state="disabled")
        self.log_box.pack(fill=tk.BOTH, expand=False)

        self._center_window(520, 480)

    def _center_window(self, width, height):
        self.update_idletasks()
        screen_w = self.winfo_screenwidth()
        screen_h = self.winfo_screenheight()
        x = int((screen_w - width) / 2)
        y = int((screen_h - height) / 2)
        self.geometry(f"{width}x{height}+{x}+{y}")

    def set_status(self, message, progress=None):
        if self._closed or not self.winfo_exists():
            return
        if message:
            self.status_label.config(text=message)
            self._append_log(message)
        if progress is not None:
            self.progress["value"] = max(0, min(100, progress))
        self.update_idletasks()

    def _append_log(self, message):
        self.log_box.configure(state="normal")
        self.log_box.insert(tk.END, "- " + message + "\n")
        self.log_box.see(tk.END)
        self.log_box.configure(state="disabled")

    def close(self):
        self._closed = True
        self.destroy()


def _get_appdata_dir():

    appdata = os.getenv("APPDATA")

    if appdata:

        return os.path.join(appdata, APP_NAME)

    return os.path.join(os.getcwd(), APP_NAME)

def _get_log_dir():
    base_dir = os.getenv("LOCALAPPDATA") or os.getenv("APPDATA") or _get_appdata_dir()
    return os.path.join(base_dir, APP_NAME, "logs")



def _setup_logging():

    log_dir = _get_log_dir()

    try:
        os.makedirs(log_dir, exist_ok=True)
    except OSError:
        log_dir = os.path.join(tempfile.gettempdir(), APP_NAME, "logs")
        os.makedirs(log_dir, exist_ok=True)

    logger = logging.getLogger("reca")

    logger.setLevel(logging.INFO)

    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    app_log = logging.handlers.RotatingFileHandler(

        os.path.join(log_dir, "app.log"), maxBytes=2 * 1024 * 1024, backupCount=3, encoding="utf-8"

    )

    app_log.setLevel(logging.INFO)

    app_log.setFormatter(formatter)

    error_log = logging.handlers.RotatingFileHandler(

        os.path.join(log_dir, "error.log"), maxBytes=2 * 1024 * 1024, backupCount=3, encoding="utf-8"

    )

    error_log.setLevel(logging.ERROR)

    error_log.setFormatter(formatter)

    logger.handlers = []

    logger.addHandler(app_log)

    logger.addHandler(error_log)

    logger.propagate = False

    return logger



LOG = _setup_logging()



def _get_env_path():

    env_override = os.getenv("RECA_ENV_PATH")

    if env_override:

        return env_override

    return os.path.join(_get_appdata_dir(), ".env")



def _load_credentials():

    env_path = _get_env_path()

    if os.path.exists(env_path):

        load_dotenv(dotenv_path=env_path, override=True)

    else:

        load_dotenv()



_load_credentials()

SUPABASE_URL = os.getenv("SUPABASE_URL")

SUPABASE_ANON_KEY = os.getenv("SUPABASE_ANON_KEY") or os.getenv("SUPABASE_KEY")
SUPABASE_AUTH_EMAIL = (os.getenv("SUPABASE_AUTH_EMAIL") or DEFAULT_SUPABASE_AUTH_EMAIL).strip()
SUPABASE_AUTH_PASSWORD = (os.getenv("SUPABASE_AUTH_PASSWORD") or DEFAULT_SUPABASE_AUTH_PASSWORD).strip()


def _ensure_credentials():
    if not SUPABASE_URL or not SUPABASE_ANON_KEY:
        LOG.error("Missing Supabase credentials")
        return False
    if not SUPABASE_AUTH_EMAIL or not SUPABASE_AUTH_PASSWORD:
        LOG.error("Missing Supabase auth credentials")
        return False
    return True


def _ensure_authenticated(client: Client):
    try:
        session = client.auth.get_session()
        if session and getattr(session, "access_token", None):
            return True

        response = client.auth.sign_in_with_password(
            {
                "email": SUPABASE_AUTH_EMAIL,
                "password": SUPABASE_AUTH_PASSWORD,
            }
        )
        session = response.session or client.auth.get_session()
        if not session or not getattr(session, "access_token", None):
            LOG.error("Supabase auth failed: missing access token for %s", SUPABASE_AUTH_EMAIL)
            return False

        LOG.info("Supabase autologin successful for %s", SUPABASE_AUTH_EMAIL)
        return True
    except Exception:
        LOG.exception("Supabase autologin failed for %s", SUPABASE_AUTH_EMAIL)
        return False


def _strip_invisible_chars(text):
    if text is None:
        return None
    return "".join(
        ch for ch in text
        if unicodedata.category(ch) not in ("Cf", "Cc") or ch in ("\n", "\t", "\r")
    )


def _clean_text(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    text = _strip_invisible_chars(text)
    text = re.sub(r"\s+", " ", text).strip()
    return text if text else None


def _normalize_nit(value):
    text = _clean_text(value)
    if not text:
        return ""
    text = text.replace(" ", "")
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    return text.lower()


def _normalize_name(value):
    text = _clean_text(value)
    return text.lower() if text else ""


def _header_key(value):
    text = _clean_text(value) or ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def _make_il_password_hash(password, iterations=260000):
    # Hash compatible with RECA Inclusion Laboral offline login verification.
    pwd = str(password or "")
    salt = secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", pwd.encode("utf-8"), salt, iterations)
    salt_b64 = base64.urlsafe_b64encode(salt).decode("ascii").rstrip("=")
    digest_b64 = base64.urlsafe_b64encode(digest).decode("ascii").rstrip("=")
    return f"pbkdf2_sha256${iterations}${salt_b64}${digest_b64}"



def _parse_version(value):

    value = (value or "").strip().lstrip("v")

    parts = value.split(".")

    try:

        return tuple(int(part) for part in parts)

    except ValueError:

        return ()



def _is_newer_version(latest):

    current = _parse_version(APP_VERSION)

    latest_version = _parse_version(latest)

    if not current or not latest_version:

        return False

    return latest_version > current



def _get_latest_release():

    url = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/releases/latest"

    response = requests.get(url, timeout=10, headers=REQUEST_HEADERS)

    if response.status_code != 200:

        LOG.error("Update check failed: %s", response.status_code)

        return None

    return response.json()

def _get_asset_url(release, asset_name):
    for asset in release.get("assets", []):
        if asset.get("name") == asset_name:
            return asset.get("browser_download_url")
    return None




def _download_asset(release):
    download_url = _get_asset_url(release, UPDATE_ASSET_NAME)
    if not download_url:
        return None
    target = os.path.join(tempfile.gettempdir(), UPDATE_ASSET_NAME)
    with requests.get(download_url, stream=True, timeout=60, headers=REQUEST_HEADERS) as resp:
        resp.raise_for_status()
        with open(target, "wb") as handler:
            for chunk in resp.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    handler.write(chunk)
    return target


def _download_hash(release):
    download_url = _get_asset_url(release, UPDATE_HASH_NAME)
    if not download_url:
        return None
    response = requests.get(download_url, timeout=10, headers=REQUEST_HEADERS)
    response.raise_for_status()
    return response.text.strip()


def _sha256_file(path):
    digest = hashlib.sha256()
    with open(path, "rb") as handler:
        for chunk in iter(lambda: handler.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _get_error_log_path():
    return os.path.join(_get_log_dir(), "error.log")


def _restart_application():
    try:
        if getattr(sys, "frozen", False):
            subprocess.Popen([sys.executable])
        else:
            subprocess.Popen([sys.executable, os.path.abspath(__file__)])
        return True
    except Exception:
        LOG.exception("Failed to restart application")
        return False




def _is_admin():
    try:
        return bool(ctypes.windll.shell32.IsUserAnAdmin())
    except Exception:
        return False


def _run_installer(installer_path):

    args = ["/SILENT", "/SUPPRESSMSGBOXES", "/NORESTART"]

    if os.name != "nt":
        try:
            subprocess.Popen([installer_path] + args)
            return True
        except Exception:
            LOG.exception("Failed to launch installer")
            return False

    try:
        current_pid = os.getpid()
        installer_ps = installer_path.replace("'", "''")
        error_log_ps = _get_error_log_path().replace("'", "''")

        if getattr(sys, "frozen", False):
            relaunch_file = sys.executable
            relaunch_args = []
        else:
            relaunch_file = sys.executable
            relaunch_args = [os.path.abspath(__file__)]

        relaunch_file_ps = relaunch_file.replace("'", "''")
        relaunch_args_ps = "@(" + ",".join("'" + arg.replace("'", "''") + "'" for arg in relaunch_args) + ")"

        helper_path = os.path.join(tempfile.gettempdir(), "reca_update_helper.ps1")
        helper_script = """param([int]$PidToWait,[string]$Installer,[string]$LogPath,[string]$Relaunch,[string[]]$RelaunchArgs)
Add-Type -AssemblyName System.Windows.Forms
$ErrorActionPreference = 'Stop'
try {
  while (Get-Process -Id $PidToWait -ErrorAction SilentlyContinue) { Start-Sleep -Milliseconds 250 }
  $proc = Start-Process -FilePath $Installer -ArgumentList @('/SILENT','/SUPPRESSMSGBOXES','/NORESTART') -Verb RunAs -Wait -PassThru
  if ($proc.ExitCode -ne 0) {
    Add-Content -Path $LogPath -Value ("{0} [ERROR] Installer exited with code {1}" -f (Get-Date -Format 'yyyy-MM-dd HH:mm:ss,fff'), $proc.ExitCode)
    [System.Windows.Forms.MessageBox]::Show("La actualizacion fallo (codigo $($proc.ExitCode)). Revisa: $LogPath", 'Actualizacion', 'OK', 'Error') | Out-Null
    exit $proc.ExitCode
  }
  Start-Process -FilePath $Relaunch -ArgumentList $RelaunchArgs | Out-Null
} catch {
  try { Add-Content -Path $LogPath -Value ("{0} [ERROR] Update helper failed: {1}" -f (Get-Date -Format 'yyyy-MM-dd HH:mm:ss,fff'), $_.Exception.Message) } catch {}
  [System.Windows.Forms.MessageBox]::Show("Ocurrio un error durante la actualizacion. Revisa: $LogPath", 'Actualizacion', 'OK', 'Error') | Out-Null
  exit 1
}
"""
        with open(helper_path, "w", encoding="utf-8") as handler:
            handler.write(helper_script)

        command = (
            "& '{helper}' -PidToWait {pid} -Installer '{installer}' -LogPath '{log}' "
            "-Relaunch '{relaunch}' -RelaunchArgs {relaunch_args}"
        ).format(
            helper=helper_path.replace("'", "''"),
            pid=current_pid,
            installer=installer_ps,
            log=error_log_ps,
            relaunch=relaunch_file_ps,
            relaunch_args=relaunch_args_ps,
        )
        subprocess.Popen(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-WindowStyle", "Hidden", "-Command", command],
            creationflags=subprocess.CREATE_NO_WINDOW,
        )
        return True
    except Exception:
        LOG.exception("Failed to launch installer helper")
        return False



def check_for_updates():

    try:

        release = _get_latest_release()

        if not release:

            return False

        latest_tag = release.get("tag_name", "")

        if not _is_newer_version(latest_tag):

            return False

        installer_path = _download_asset(release)
        if not installer_path:
            LOG.error("Update asset not found in release %s", latest_tag)
            return False
        expected_hash = _download_hash(release)
        if not expected_hash:
            LOG.error("Update hash not found in release %s", latest_tag)
            return False
        expected_value = expected_hash.split()[0].lower()
        actual_value = _sha256_file(installer_path).lower()
        if expected_value != actual_value:
            LOG.error("Update hash mismatch for %s", latest_tag)
            return False
        LOG.info("Update verified for version %s", latest_tag)
        if _run_installer(installer_path):
            messagebox.showinfo(
                "Actualización",
                "Se iniciará la instalación. La aplicación se cerrará y se abrirá al finalizar.",
            )
            return True
        LOG.error("Update installer failed for %s", latest_tag)
        messagebox.showerror(
            "Actualización",
            "No se pudo completar la actualización.\n"
            f"Revisa los logs en {_get_error_log_path()}",
        )
        return False

    except Exception:

        LOG.exception("Auto-update failed")
        messagebox.showerror(
            "Actualización",
            "Ocurrió un error al intentar actualizar.\n"
            f"Revisa los logs en {_get_error_log_path()}",
        )
        return False





def conectar_supabase():
    """
    Establece conexion con Supabase

    Returns:
        Client: Cliente de Supabase o None si hay error
    """
    if not _ensure_credentials():
        messagebox.showerror("Error", "Credenciales no configuradas")
        return None
    try:
        client = create_client(SUPABASE_URL, SUPABASE_ANON_KEY)
        if not _ensure_authenticated(client):
            messagebox.showerror(
                "Error",
                "No se pudo iniciar sesion automatica en Supabase.\n"
                f"Revisa los logs en {_get_error_log_path()}",
            )
            return None
        return client
    except Exception as e:
        LOG.exception("Error conectando a Supabase")
        messagebox.showerror("Error", f"Error conectando: {e}")
        return None

# ============================================

# VENTANA DE FORMULARIO

# ============================================



class FormularioEmpresa(tk.Toplevel):
    """Ventana modal para crear o editar empresas"""

    LARGE_TEXT_FIELDS = {"cargo", "contacto_empresa", "telefono_empresa", "correo_1", "observaciones"}

    # Configuracion de campos del formulario
    FIELD_CONFIG = {
        "nombre_empresa": ("Nombre Empresa *", True),
        "nit_empresa": ("NIT", False),
        "direccion_empresa": ("Dirección", False),
        "ciudad_empresa": ("Ciudad", False),
        "correo_1": ("Email(s)", False),
        "contacto_empresa": ("Contacto(s)", False),
        "cargo": ("Cargo", False),
        "telefono_empresa": ("Teléfono(s)", False),
        "sede_empresa": ("Sede Empresa", False),
        "zona_empresa": ("Zona Compensar", False),
        "responsable_visita": ("Responsable Visita", False),
        "asesor": ("Asesor", False),
        "correo_asesor": ("Email Asesor", False),
        "profesional_asignado": ("Profesional Asignado", False),
        "correo_profesional": ("Email Profesional", False),
        "caja_compensacion": ("Caja Compensación", False),
        "estado": ("Estado *", True),
        "observaciones": ("Observaciones", False),
    }

    SECCIONES = [
        ("Empresa", [
            "nombre_empresa",
            "nit_empresa",
            "direccion_empresa",
            "ciudad_empresa",
            "sede_empresa",
            "responsable_visita",
            "cargo",
            "contacto_empresa",
            "telefono_empresa",
            "correo_1",
            "estado",
        ], "#E6F4EA"),
        ("Compensar", [
            "caja_compensacion",
            "zona_empresa",
            "asesor",
            "correo_asesor",
        ], "#FFF3E0"),
        ("RECA", [
            "profesional_asignado",
            "correo_profesional",
        ], "#F3E5F5"),
        ("Observaciones", [
            "observaciones",
        ], "#F7F5FA"),
    ]

    ESTADOS_DISPONIBLES = ["Activa", "En Proceso", "Pausada", "Cerrada", "Inactiva"]
    CAJAS_COMPENSACION = ["Compensar", "No Compensar"]
    ZONAS_COMPENSAR = {
        "soacha": "Soacha",
        "universidad_compensar": "Universidad Compensar",
        "kennedy": "Kennedy",
        "chapinero": "Chapinero",
        "bosa": "Bosa",
        "mosquera": "Mosquera",
        "cajica": "Cajicá",
        "girardot": "Girardot",
        "suba": "Suba",
        "empleabilidad_estrategica": "Empleabilidad Estrategica",
        "por_confirmar": "Por Confirmar",
    }

    def __init__(self, parent, supabase, empresa=None):
        """
        Inicializa el formulario

        Args:
            parent: Ventana padre
            supabase: Cliente de Supabase
            empresa: Datos de empresa a editar (None para crear nueva)
        """
        super().__init__(parent)
        self.parent = parent
        self.supabase = supabase
        self.empresa = empresa
        self.resultado = None
        self.campos = {}
        self._asesores = []
        self._asesores_correo = {}
        self._profesionales = []
        self._profesionales_correo = {}

        # Configurar ventana
        titulo = "Editar Empresa" if empresa else "Nueva Empresa"
        self.title(titulo)
        self.geometry("900x700")
        _maximize_window(self)

        # Cargar catalogos
        self._cargar_catalogos()

        # Crear interfaz
        self.crear_formulario()

        # Convertir en modal
        self.transient(parent)
        self.grab_set()
        self.bind("<Control-s>", lambda e: self.guardar())
        self.bind("<Escape>", lambda e: self.destroy())

    def crear_formulario(self):

        """Construye la interfaz del formulario con un grid scrollable."""

        scroll = ScrollableFrame(self, bg=COLOR_LIGHT_BG)
        scroll.pack(fill="both", expand=True)
        scroll.content.grid_columnconfigure(0, weight=1)

        titulo_text = "Editar Empresa" if self.empresa else "Nueva Empresa"

        titulo_label = tk.Label(

            scroll.content,

            text=titulo_text,

            font=FONT_H2,
            bg=COLOR_LIGHT_BG,
            fg=COLOR_PURPLE,

        )

        titulo_label.grid(row=0, column=0, columnspan=4, pady=20)

        last_row = self._crear_campos(scroll.content)

        self._crear_botones(scroll.content, last_row + 1)



    def _crear_campos(self, parent):

        """

        Crea todos los campos del formulario agrupados por secciones.

        Args:

            parent: Frame contenedor

        """

        row = 1
        for titulo, campos, bg in self.SECCIONES:
            title_label = tk.Label(
                parent,
                text=titulo,
                bg=bg,
                fg=COLOR_PURPLE,
                font=FONT_H3,
            )
            frame = tk.LabelFrame(
                parent,
                labelwidget=title_label,
                bg=bg,
                padx=SP_MD,
                pady=SP_SM,
                bd=1,
                relief="flat",
            )
            frame.grid(row=row, column=0, columnspan=4, sticky="ew", padx=12, pady=8)
            frame.grid_columnconfigure(1, weight=1)
            row += 1

            inner_row = 0
            for campo in campos:
                label_text, _required = self.FIELD_CONFIG.get(campo, (campo, False))
                tk.Label(
                    frame,
                    text=label_text,
                    font=FONT_BODY_BOLD,
                    bg=bg,
                ).grid(row=inner_row, column=0, sticky="w", padx=6, pady=4)

                if campo == "estado":
                    widget = self._crear_campo_estado(frame)
                elif campo == "caja_compensacion":
                    widget = self._crear_campo_caja_compensacion(frame)
                elif campo == "zona_empresa":
                    widget = self._crear_campo_zona_compensar(frame)
                elif campo in self.LARGE_TEXT_FIELDS:
                    widget = self._crear_campo_texto_largo(frame, campo)
                elif campo == "asesor":
                    widget = self._crear_campo_asesor(frame)
                elif campo == "profesional_asignado":
                    widget = self._crear_campo_profesional(frame)
                else:
                    widget = self._crear_campo_texto(frame, campo)

                self.campos[campo] = widget
                if campo in self.LARGE_TEXT_FIELDS:
                    widget.grid(row=inner_row, column=1, columnspan=3, sticky="ew", padx=6, pady=4)
                else:
                    widget.grid(row=inner_row, column=1, columnspan=3, sticky="w", padx=6, pady=4)
                inner_row += 1

        return row



    def _crear_campo_estado(self, parent):

        """Crea combobox para el campo estado"""

        widget = ttk.Combobox(

            parent,

            values=self.ESTADOS_DISPONIBLES,

            state="readonly",

            width=40,

            font=FONT_BODY

        )



        # Establecer valor inicial

        if self.empresa:

            valor = self.empresa.get("estado", "En Proceso") or "En Proceso"

            widget.set(valor)

        else:

            widget.set("En Proceso")



        return widget

    def _crear_campo_texto_largo(self, parent, campo):

        """Crea areas de texto amplias para campos con contenido largo"""

        widget = scrolledtext.ScrolledText(

            parent,

            width=50,

            height=5 if campo == "observaciones" else 3,

            wrap=tk.WORD,

            font=FONT_BODY

        )



        if self.empresa:

            valor = self.empresa.get(campo, "") or ""

            widget.insert("1.0", valor)



        return widget



    def _crear_campo_observaciones(self, parent):

        """Crea área de texto para observaciones"""

        widget = scrolledtext.ScrolledText(

            parent,

            width=50,

            height=5,

            font=FONT_BODY

        )



        # Cargar valor si existe

        if self.empresa:

            valor = self.empresa.get("observaciones", "") or ""

            widget.insert("1.0", valor)



        return widget

    def _crear_campo_caja_compensacion(self, parent):
        """Crea combobox para el campo caja de compensacion"""
        widget = ttk.Combobox(
            parent,
            values=self.CAJAS_COMPENSACION,
            state="readonly",
            width=40,
            font=FONT_BODY
        )

        if self.empresa:
            valor = self.empresa.get("caja_compensacion", "Compensar") or "Compensar"
            widget.set(valor)
        else:
            widget.set("Compensar")

        return widget

    def _crear_campo_cargo(self, parent):
        """Crea un area de texto amplia para cargos largos."""
        widget = scrolledtext.ScrolledText(
            parent,
            width=50,
            height=3,
            wrap=tk.WORD,
            font=FONT_BODY
        )

        if self.empresa:
            valor = self.empresa.get("cargo", "") or ""
            widget.insert("1.0", valor)

        return widget

    def _crear_campo_zona_compensar(self, parent):
        """Crea combobox editable para el campo zona_compensar usando opciones locales"""
        widget = ttk.Combobox(
            parent,
            values=list(self.ZONAS_COMPENSAR.values()),
            state="normal",
            width=40,
            font=FONT_BODY
        )

        if self.empresa:
            valor = self.empresa.get("zona_empresa", "") or ""
            widget.set(valor)

        return widget

    def _crear_campo_asesor(self, parent):
        """Crea combobox para el campo asesor"""
        widget = ttk.Combobox(
            parent,
            values=self._asesores,
            state="readonly",
            width=40,
            font=FONT_BODY
        )
        if self.empresa:
            valor = self.empresa.get("asesor", "") or ""
            widget.set(valor)
            self._actualizar_correo_asesor(valor)
        widget.bind("<<ComboboxSelected>>", lambda e: self._actualizar_correo_asesor(widget.get()))
        return widget

    def _crear_campo_profesional(self, parent):
        """Crea combobox para el campo profesional asignado"""
        widget = ttk.Combobox(
            parent,
            values=self._profesionales,
            state="normal",
            width=40,
            font=FONT_BODY
        )
        if self.empresa:
            valor = self.empresa.get("profesional_asignado", "") or ""
            widget.set(valor)
            self._actualizar_correo_profesional(valor)
        widget.bind("<<ComboboxSelected>>", lambda e: self._actualizar_correo_profesional(widget.get()))
        widget.bind("<FocusOut>", lambda e: self._actualizar_correo_profesional(widget.get()))
        return widget


    def _crear_campo_texto(self, parent, campo):

        """Crea campo de entrada de texto simple"""

        widget = tk.Entry(parent, width=60, font=FONT_BODY, highlightcolor=COLOR_TEAL, highlightthickness=1)



        # Cargar valor si existe

        if self.empresa:

            valor = self.empresa.get(campo, "") or ""

            widget.insert(0, valor)

        if campo == "correo_asesor" and not widget.get():
            asesor_widget = self.campos.get("asesor")
            if asesor_widget:
                self._actualizar_correo_asesor(asesor_widget.get())

        if campo == "correo_profesional" and not widget.get():
            profesional_widget = self.campos.get("profesional_asignado")
            if profesional_widget:
                self._actualizar_correo_profesional(profesional_widget.get())



        return widget

    def _actualizar_correo_asesor(self, nombre):
        correo = self._asesores_correo.get(nombre, "")
        widget = self.campos.get("correo_asesor")
        if widget:
            widget.delete(0, tk.END)
            widget.insert(0, correo)

    def _actualizar_correo_profesional(self, nombre):
        correo = self._profesionales_correo.get((nombre or "").strip(), "")
        widget = self.campos.get("correo_profesional")
        if widget and correo:
            widget.delete(0, tk.END)
            widget.insert(0, correo)

    def _asegurar_profesional(self, nombre, correo):
        nombre = (nombre or "").strip()
        correo = (correo or "").strip()
        if not nombre:
            return correo
        try:
            response = (self.supabase.table("profesionales")
                        .select("nombre_profesional, correo_profesional")
                        .eq("nombre_profesional", nombre)
                        .limit(1)
                        .execute())
            existente = (response.data or [])
            if existente:
                correo_existente = (existente[0].get("correo_profesional") or "").strip()
                final_correo = correo_existente or correo
                if correo and correo != correo_existente:
                    (self.supabase.table("profesionales")
                        .update({"correo_profesional": correo})
                        .eq("nombre_profesional", nombre)
                        .execute())
                    final_correo = correo
                self._profesionales_correo[nombre] = final_correo
                if nombre not in self._profesionales:
                    self._profesionales.append(nombre)
                    self._profesionales.sort()
                return final_correo

            payload = {"nombre_profesional": nombre}
            if correo:
                payload["correo_profesional"] = correo
            self.supabase.table("profesionales").insert(payload).execute()
            self._profesionales_correo[nombre] = correo
            if nombre not in self._profesionales:
                self._profesionales.append(nombre)
                self._profesionales.sort()
            return correo
        except Exception:
            LOG.exception("Error sincronizando profesional: %s", nombre)
            raise

    def _cargar_catalogos(self):
        if not self.supabase:
            return
        try:
            asesores = (self.supabase.table("asesores")
                        .select("nombre, email")
                        .order("nombre", desc=False)
                        .execute()).data or []
            nombres_asesores = []
            seen = set()
            for row in asesores:
                nombre = (row.get("nombre") or "").strip()
                if nombre and nombre not in seen:
                    seen.add(nombre)
                    nombres_asesores.append(nombre)
            self._asesores = nombres_asesores
            self._asesores_correo = {
                row.get("nombre", "").strip(): (row.get("email", "") or "").strip()
                for row in asesores if row.get("nombre")
            }
        except Exception:
            LOG.exception("Error cargando asesores")

        try:
            profesionales = (self.supabase.table("profesionales")
                             .select("nombre_profesional, correo_profesional")
                             .order("nombre_profesional", desc=False)
                             .execute()).data or []
            nombres_profesionales = []
            seen = set()
            for row in profesionales:
                nombre = (row.get("nombre_profesional") or "").strip()
                if nombre and nombre not in seen:
                    seen.add(nombre)
                    nombres_profesionales.append(nombre)
            self._profesionales = nombres_profesionales
            self._profesionales_correo = {
                row.get("nombre_profesional", "").strip(): (row.get("correo_profesional", "") or "").strip()
                for row in profesionales if row.get("nombre_profesional")
            }
        except Exception:
            LOG.exception("Error cargando profesionales")



    def _crear_botones(self, parent, row):

        """

        Crea los botones de acción del formulario



        Args:

            parent: Frame contenedor

            row: Fila donde colocar los botones

        """

        btn_frame = tk.Frame(parent, bg=COLOR_LIGHT_BG)
        btn_frame.grid(row=row, column=0, columnspan=4, pady=SP_LG)

        if self.empresa:
            _make_button(btn_frame, "Guardar Cambios", self.guardar, style="success").pack(side=tk.LEFT, padx=SP_SM)
            _make_button(btn_frame, "Eliminar", self.eliminar, style="danger").pack(side=tk.RIGHT, padx=SP_SM)
        else:
            _make_button(btn_frame, "Crear Empresa", self.guardar, style="success").pack(side=tk.LEFT, padx=SP_SM)

        _make_button(btn_frame, "Cancelar", self.destroy, style="outline").pack(side=tk.LEFT, padx=SP_SM)



    def guardar(self):

        """Guarda o actualiza la empresa en la base de datos"""



        # Recoger datos de todos los campos

        datos = {}

        for campo, widget in self.campos.items():

            if isinstance(widget, scrolledtext.ScrolledText):

                datos[campo] = widget.get("1.0", tk.END).strip()

            elif isinstance(widget, ttk.Combobox):

                datos[campo] = widget.get()

            else:

                datos[campo] = widget.get().strip()



        # Validar campo obligatorio

        if not datos["nombre_empresa"]:

            messagebox.showerror("Error", "El nombre de la empresa es obligatorio")

            return

        # Sincronizar profesional asignado con tabla 'profesionales'
        profesional_nombre = (datos.get("profesional_asignado") or "").strip()
        profesional_correo = (datos.get("correo_profesional") or "").strip()
        if profesional_nombre:
            datos["correo_profesional"] = self._asegurar_profesional(
                profesional_nombre,
                profesional_correo,
            )



        try:

            if self.empresa:

                # Actualizar empresa existente

                self.supabase.table("empresas").update(datos).eq("id", self.empresa["id"]).execute()
                LOG.info("Empresa actualizada: %s", datos.get("nombre_empresa"))

                messagebox.showinfo("Éxito", "Empresa actualizada correctamente")

            else:

                # Crear nueva empresa

                self.supabase.table("empresas").insert(datos).execute()
                LOG.info("Empresa creada: %s", datos.get("nombre_empresa"))

                messagebox.showinfo("Éxito", "Empresa creada correctamente")



            self.resultado = "guardado"

            self.destroy()



        except Exception as e:
            LOG.exception("Error en formulario de empresa")

            messagebox.showerror("Error", f"Error guardando empresa: {e}")



    def eliminar(self):

        """Elimina la empresa de la base de datos"""



        if not self.empresa:

            return



        # Confirmar eliminación

        confirmar = messagebox.askyesno(

            "Confirmar eliminación",

            f"¿Estás seguro de eliminar la empresa '{self.empresa['nombre_empresa']}'?\n\n"

            "Esta acción no se puede deshacer."

        )



        if not confirmar:

            return



        try:

            self.supabase.table("empresas").delete().eq("id", self.empresa["id"]).execute()
            LOG.info("Empresa eliminada: %s", self.empresa.get("nombre_empresa"))

            messagebox.showinfo("Éxito", "Empresa eliminada correctamente")

            self.resultado = "eliminado"

            self.destroy()

        except Exception as e:
            LOG.exception("Error en formulario de empresa")

            messagebox.showerror("Error", f"Error eliminando empresa: {e}")



# ============================================
# FORMULARIOS GENERICOS
# ============================================


class FormularioEntidad(tk.Toplevel):
    """Ventana modal para crear o editar registros genericos"""

    def __init__(self, parent, supabase, tabla, campos, key_field, registro=None, titulo="Registro"):
        super().__init__(parent)
        self.parent = parent
        self.supabase = supabase
        self.tabla = tabla
        self.campos_config = campos
        self.key_field = key_field
        self.registro = registro
        self.resultado = None
        self.widgets = {}
        self._original_key = registro.get(key_field) if registro else None

        self.title(titulo)
        self.geometry("700x600")
        _maximize_window(self)

        self._crear_formulario()

        self.transient(parent)
        self.grab_set()
        self.bind("<Control-s>", lambda e: self.guardar())
        self.bind("<Escape>", lambda e: self.destroy())

    def _crear_formulario(self):
        scroll = ScrollableFrame(self)
        scroll.pack(fill=tk.BOTH, expand=True, padx=16, pady=16)
        container = scroll.content
        container.grid_columnconfigure(1, weight=1)

        row = 0
        for campo, label, required, widget_type in self.campos_config:
            tk.Label(
                container,
                text=label,
                font=FONT_BODY_BOLD,
            ).grid(row=row, column=0, sticky="w", padx=6, pady=4)

            if widget_type == "text":
                widget = scrolledtext.ScrolledText(container, width=50, height=4, font=FONT_BODY)
                if self.registro:
                    valor = self.registro.get(campo, "") or ""
                    widget.insert("1.0", valor)
                widget.grid(row=row, column=1, sticky="ew", padx=6, pady=4)
            else:
                widget = tk.Entry(container, width=50, font=FONT_BODY, highlightcolor=COLOR_TEAL, highlightthickness=1)
                if self.registro:
                    valor = self.registro.get(campo, "") or ""
                    widget.insert(0, valor)
                widget.grid(row=row, column=1, sticky="w", padx=6, pady=4)

            self.widgets[campo] = widget
            row += 1

        btn_frame = tk.Frame(container, bg=COLOR_LIGHT_BG)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=SP_MD)

        if self.registro:
            _make_button(btn_frame, "Guardar Cambios", self.guardar, style="success").pack(side=tk.LEFT, padx=SP_SM)
            _make_button(btn_frame, "Eliminar", self.eliminar, style="danger").pack(side=tk.RIGHT, padx=SP_SM)
        else:
            _make_button(btn_frame, "Crear", self.guardar, style="success").pack(side=tk.LEFT, padx=SP_SM)

        _make_button(btn_frame, "Cancelar", self.destroy, style="outline").pack(side=tk.LEFT, padx=SP_SM)

    def guardar(self):
        datos = {}
        for campo, label, required, widget_type in self.campos_config:
            if widget_type == "text":
                valor = self.widgets[campo].get("1.0", tk.END).strip()
            else:
                valor = self.widgets[campo].get().strip()
            if required and not valor:
                messagebox.showerror("Error", f"El campo '{label}' es obligatorio")
                return
            datos[campo] = valor

        try:
            if self.registro:
                self.supabase.table(self.tabla).update(datos).eq(self.key_field, self._original_key).execute()
                messagebox.showinfo("Éxito", "Registro actualizado correctamente")
            else:
                self.supabase.table(self.tabla).insert(datos).execute()
                messagebox.showinfo("Éxito", "Registro creado correctamente")
            self.resultado = "guardado"
            self.destroy()
        except Exception as e:
            LOG.exception("Error guardando registro en %s", self.tabla)
            messagebox.showerror("Error", f"Error guardando registro: {e}")

    def eliminar(self):
        if not self.registro:
            return
        confirmar = messagebox.askyesno("Confirmar eliminación", "¿Eliminar este registro?")
        if not confirmar:
            return
        try:
            self.supabase.table(self.tabla).delete().eq(self.key_field, self._original_key).execute()
            messagebox.showinfo("Éxito", "Registro eliminado correctamente")
            self.resultado = "eliminado"
            self.destroy()
        except Exception as e:
            LOG.exception("Error eliminando registro en %s", self.tabla)
            messagebox.showerror("Error", f"Error eliminando registro: {e}")


# ============================================
# APPS GENERICAS
# ============================================


class AppEntidad:
    """Aplicacion generica para gestionar tablas simples"""

    def __init__(self, root, table, columns, labels, widths, form_config, key_field, title, order_by):
        self.root = root
        self.root.title(title)
        self.root.geometry("1200x720")
        _maximize_window(self.root)

        self.table = table
        self.columns = columns
        self.labels = labels
        self.widths = widths
        self.form_config = form_config
        self.key_field = key_field
        self.order_by = order_by

        self.supabase = conectar_supabase()
        self.registros = []
        self._registro_por_key = {}
        self.registro_seleccionado = None

        self._crear_interfaz()
        self.cargar_registros()

    def _crear_interfaz(self):
        header = tk.Frame(self.root, bg=COLOR_PURPLE, height=72)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(
            header,
            text=self.root.title(),
            font=FONT_H1,
            bg=COLOR_PURPLE,
            fg="white",
        ).pack(side=tk.LEFT, padx=SP_MD)

        tabla_frame = tk.Frame(self.root)
        tabla_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        self.contador_label = tk.Label(
            tabla_frame,
            text="Resultados: 0 registros",
            font=FONT_BODY_BOLD
        )
        self.contador_label.grid(row=0, column=0, sticky="w", pady=5)

        self.tree = ttk.Treeview(
            tabla_frame,
            columns=self.columns,
            show="headings",
            height=18
        )
        for col in self.columns:
            self.tree.heading(col, text=self.labels.get(col, col))
            self.tree.column(col, width=self.widths.get(col, 120))
        scroll_y = ttk.Scrollbar(tabla_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scroll_y.set)
        self.tree.grid(row=1, column=0, sticky="nsew")
        scroll_y.grid(row=1, column=1, sticky="ns")
        tabla_frame.grid_rowconfigure(1, weight=1)
        tabla_frame.grid_columnconfigure(0, weight=1)

        self.tree.bind("<ButtonRelease-1>", self._seleccionar)
        self.tree.bind("<Double-1>", self._abrir_editar)

        # Filas alternadas
        self.tree.tag_configure("oddrow", background=TREE_ROW_ALT)
        self.tree.tag_configure("evenrow", background=COLOR_WHITE)

        btn_frame = tk.Frame(self.root, bg=COLOR_LIGHT_BG)
        btn_frame.pack(fill=tk.X, padx=SP_LG, pady=SP_SM)

        _make_button(btn_frame, "Nuevo", self.nuevo_registro, style="primary").pack(side=tk.LEFT, padx=SP_XS)
        _make_button(btn_frame, "Editar", self.editar_registro, style="secondary").pack(side=tk.LEFT, padx=SP_XS)
        _make_button(btn_frame, "Refrescar", self.cargar_registros, style="outline").pack(side=tk.LEFT, padx=SP_XS)

        if self.table == "profesionales":
            _make_button(btn_frame, "Restablecer contraseña", self.restablecer_contrasena_profesional, style="warning").pack(side=tk.RIGHT, padx=SP_XS)

        _make_button(btn_frame, "Eliminar", self.eliminar_registro, style="danger").pack(side=tk.RIGHT, padx=SP_XS)

        # Atajos de teclado
        self.root.bind("<Control-n>", lambda e: self.nuevo_registro())
        self.root.bind("<F5>", lambda e: self.cargar_registros())

    def restablecer_contrasena_profesional(self):
        if self.table != "profesionales":
            return
        if not self.registro_seleccionado:
            messagebox.showwarning("Aviso", "Selecciona un profesional primero")
            return
        profesional_id = self.registro_seleccionado.get("id")
        if profesional_id is None:
            messagebox.showerror("Error", "El profesional seleccionado no tiene ID valido")
            return

        nombre = (
            self.registro_seleccionado.get("nombre_profesional")
            or self.registro_seleccionado.get("usuario_login")
            or str(profesional_id)
        )
        usuario_login = (self.registro_seleccionado.get("usuario_login") or "").strip()
        correo_profesional = (self.registro_seleccionado.get("correo_profesional") or "").strip()
        nueva = DEFAULT_PROFESIONAL_TEMP_PASSWORD

        if not usuario_login:
            messagebox.showerror(
                "Error",
                "Este profesional no tiene 'Usuario Login' configurado.\n"
                "Sin ese dato no puede ingresar a RECA Inclusion Laboral.",
            )
            return
        if not correo_profesional:
            messagebox.showerror(
                "Error",
                "Este profesional no tiene 'Correo Profesional' configurado.\n"
                "Sin ese dato no puede iniciar sesion en RECA Inclusion Laboral.",
            )
            return
        confirmar = messagebox.askyesno(
            "Restablecer contrasena",
            "Se restablecera la contrasena temporal para RECA Inclusion Laboral.\n\n"
            f"Profesional: {nombre}\n"
            f"Usuario login: {usuario_login}\n"
            f"Correo Auth: {correo_profesional}\n"
            f"Contrasena temporal: {nueva}\n\n"
            "Deseas continuar?",
            parent=self.root,
        )
        if not confirmar:
            return
        try:
            hash_value = _make_il_password_hash(nueva)
            response = self.supabase.rpc(
                "admin_reset_profesional_password",
                {
                    "p_profesional_id": int(profesional_id),
                    "p_new_password": nueva,
                    "p_new_password_hash": hash_value,
                },
            ).execute()

            result = response.data
            auth_updated = False
            has_auth_user = False
            if isinstance(result, dict):
                auth_updated = bool(result.get("auth_user_updated"))
                has_auth_user = bool(result.get("has_auth_user"))
            LOG.info(
                "Password reset for profesional_id=%s auth_updated=%s has_auth_user=%s usuario_login=%s",
                profesional_id,
                auth_updated,
                has_auth_user,
                usuario_login,
            )

            messagebox.showinfo(
                "Contrasena restablecida",
                "Contrasena actualizada correctamente.\n\n"
                f"Usuario login: {usuario_login}\n"
                f"Contrasena temporal: {nueva}\n"
                f"Cuenta Auth enlazada: {'Si' if has_auth_user else 'No'}\n"
                f"Actualizada en Auth: {'Si' if auth_updated else 'No'}",
            )
            self.cargar_registros()
        except Exception as e:
            LOG.exception("Error restableciendo contrasena de profesional id=%s", profesional_id)
            messagebox.showerror(
                "Error",
                "No se pudo restablecer la contrasena.\n"
                f"Detalle: {e}\n"
                f"Revisa logs en {_get_error_log_path()}",
            )

    def cargar_registros(self):
        if not self.supabase:
            return
        try:
            query = self.supabase.table(self.table).select("*")
            if self.order_by:
                query = query.order(self.order_by, desc=False)
            data = query.execute().data or []
            self.registros = data
            self._registro_por_key = {}
            self._limpiar_tabla()
            self._mostrar_registros(data)
            self.contador_label.config(text=f"Resultados: {len(data)} registros")
        except Exception as e:
            LOG.exception("Error cargando registros de %s", self.table)
            messagebox.showerror("Error", f"Error cargando registros: {e}")

    def _limpiar_tabla(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

    def _mostrar_registros(self, registros):
        for idx, registro in enumerate(registros):
            key_value = registro.get(self.key_field)
            key_str = str(key_value) if key_value is not None else ""
            if key_str:
                self._registro_por_key[key_str] = registro
            row_tag = "oddrow" if idx % 2 else "evenrow"
            tags = (key_str, row_tag) if key_str else (row_tag,)
            values = [registro.get(col, "") for col in self.columns]
            self.tree.insert("", tk.END, values=values, tags=tags)

    def _seleccionar(self, event):
        selection = self.tree.selection()
        if selection:
            item = self.tree.item(selection[0])
            registro_key = item["tags"][0] if item["tags"] else None
            if registro_key is not None:
                registro_key = str(registro_key)
            self.registro_seleccionado = self._registro_por_key.get(registro_key)

    def _abrir_editar(self, event=None):
        if not self.registro_seleccionado:
            messagebox.showwarning("Aviso", "Selecciona un registro primero")
            return
        ventana = FormularioEntidad(
            self.root,
            self.supabase,
            self.table,
            self.form_config,
            self.key_field,
            registro=self.registro_seleccionado,
            titulo=f"Editar {self.table}",
        )
        self.root.wait_window(ventana)
        if ventana.resultado:
            self.cargar_registros()

    def nuevo_registro(self):
        if not self.supabase:
            return
        ventana = FormularioEntidad(
            self.root,
            self.supabase,
            self.table,
            self.form_config,
            self.key_field,
            registro=None,
            titulo=f"Nuevo {self.table}",
        )
        self.root.wait_window(ventana)
        if ventana.resultado:
            self.cargar_registros()

    def editar_registro(self):
        if not self.registro_seleccionado:
            messagebox.showwarning("Aviso", "Selecciona un registro primero")
            return
        self._abrir_editar()

    def eliminar_registro(self):
        if not self.registro_seleccionado:
            messagebox.showwarning("Aviso", "Selecciona un registro primero")
            return
        confirmar = messagebox.askyesno("Confirmar", "¿Eliminar el registro seleccionado?")
        if not confirmar:
            return
        try:
            self.supabase.table(self.table).delete().eq(
                self.key_field, self.registro_seleccionado.get(self.key_field)
            ).execute()
            messagebox.showinfo("Éxito", "Registro eliminado")
            self.cargar_registros()
        except Exception as e:
            LOG.exception("Error eliminando registro en %s", self.table)
            messagebox.showerror("Error", f"Error eliminando registro: {e}")


class AppMenu:
    """Pantalla principal de opciones"""

    def __init__(self, root):
        self.root = root
        self.root.title("RECA - Panel principal")
        self.root.geometry("1000x700")
        _maximize_window(self.root)
        self._latest_version = None

        self._crear_interfaz()
        self._fetch_latest_version_async()

    def _crear_interfaz(self):
        # Header
        header = tk.Frame(self.root, bg=COLOR_PURPLE, height=72)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        tk.Label(
            header,
            text="RECA - Panel principal",
            font=FONT_H1,
            bg=COLOR_PURPLE,
            fg="white",
        ).pack(side=tk.LEFT, padx=SP_LG)

        self.update_btn = _make_button(header, "Actualizar app", self._manual_update, style="outline", font=FONT_SMALL)
        self.update_btn.pack(side=tk.RIGHT, padx=SP_LG)

        # Body
        body_scroll = ScrollableFrame(self.root, bg=COLOR_LIGHT_BG)
        body_scroll.pack(fill=tk.BOTH, expand=True)
        body = body_scroll.content

        tk.Label(
            body,
            text="Selecciona un modulo",
            font=FONT_H2,
            bg=COLOR_LIGHT_BG,
            fg=COLOR_PURPLE,
        ).pack(pady=(SP_LG * 2, SP_MD))

        # Cards grid
        cards_frame = tk.Frame(body, bg=COLOR_LIGHT_BG)
        cards_frame.pack(pady=SP_MD)

        modulos = [
            ("Empresas", "Gestion de empresas y clientes", self._abrir_empresas),
            ("Asesores", "Asesores de Compensar", self._abrir_asesores),
            ("Gestores", "Gestores de empleo", self._abrir_gestores),
            ("Profesionales", "Profesionales de RECA", self._abrir_profesionales),
            ("Interpretes", "Interpretes de lengua de senas", self._abrir_interpretes),
        ]

        for idx, (titulo, desc, cmd) in enumerate(modulos):
            row, col = divmod(idx, 2)
            self._crear_card_modulo(cards_frame, titulo, desc, cmd).grid(
                row=row, column=col, padx=SP_SM, pady=SP_SM, sticky="nsew"
            )

        cards_frame.grid_columnconfigure(0, weight=1)
        cards_frame.grid_columnconfigure(1, weight=1)

        self._crear_footer()

    def _crear_footer(self):
        footer = tk.Frame(self.root, bg=COLOR_LIGHT_BG, height=48)
        footer.pack(fill=tk.X, side=tk.BOTTOM)
        footer.pack_propagate(False)

        info_wrap = tk.Frame(footer, bg=COLOR_LIGHT_BG)
        info_wrap.pack(side=tk.LEFT, anchor="sw", padx=8, pady=4)

        self.version_label = tk.Label(
            info_wrap,
            text=f"Version: {APP_VERSION} | GitHub: --",
            font=FONT_SMALL,
            bg=COLOR_LIGHT_BG,
            fg="#666666",
            justify="left",
            anchor="w",
        )
        self.version_label.pack(anchor="w")

    def _fetch_latest_version_async(self):
        def worker():
            latest = None
            try:
                release = _get_latest_release()
                if release:
                    latest = release.get("tag_name", "") or None
            except Exception:
                LOG.exception("Error fetching latest release for menu footer")
            self.root.after(0, lambda: self._update_version_label(latest))

        threading.Thread(target=worker, daemon=True).start()

    def _update_version_label(self, latest):
        self._latest_version = latest
        latest_text = latest if latest else "--"
        if self.version_label.winfo_exists():
            self.version_label.config(text=f"Version: {APP_VERSION} | GitHub: {latest_text}")

    def _manual_update(self):
        latest = self._latest_version
        if not latest:
            try:
                release = _get_latest_release()
                latest = release.get("tag_name", "") if release else ""
                self._update_version_label(latest or None)
            except Exception:
                latest = ""
        if not _is_newer_version(latest or ""):
            messagebox.showinfo("Actualización", "Ya tienes la última versión instalada.")
            return
        if check_for_updates():
            self.root.destroy()

    def _crear_card_modulo(self, parent, titulo, descripcion, comando):
        card = tk.Frame(
            parent,
            bg=COLOR_WHITE,
            highlightthickness=1,
            highlightbackground=COLOR_BORDER,
            cursor="hand2",
            padx=SP_LG,
            pady=SP_MD,
        )
        card.configure(width=280, height=90)
        card.pack_propagate(False)

        tk.Label(
            card, text=titulo, font=FONT_H3, bg=COLOR_WHITE, fg=COLOR_PURPLE, anchor="w",
        ).pack(fill=tk.X, pady=(SP_XS, 2))
        tk.Label(
            card, text=descripcion, font=FONT_BODY, bg=COLOR_WHITE, fg="#666666", anchor="w",
        ).pack(fill=tk.X)

        def on_enter(_e):
            card.config(highlightbackground=COLOR_TEAL, highlightthickness=2)
        def on_leave(_e):
            card.config(highlightbackground=COLOR_BORDER, highlightthickness=1)
        def on_click(_e):
            comando()

        for widget in (card, *card.winfo_children()):
            widget.bind("<Enter>", on_enter)
            widget.bind("<Leave>", on_leave)
            widget.bind("<Button-1>", on_click)
            try:
                widget.config(cursor="hand2")
            except tk.TclError:
                pass

        return card

    def _crear_ventana(self, titulo):
        ventana = tk.Toplevel(self.root)
        ventana.title(titulo)
        if os.path.exists(LOGO_PATH):
            try:
                icon = tk.PhotoImage(file=LOGO_PATH)
                ventana.iconphoto(True, icon)
                ventana._window_icon = icon
            except Exception:
                pass
        return ventana

    def _abrir_empresas(self):
        ventana = self._crear_ventana("RECA - Empresas")
        AppRECA(ventana)

    def _abrir_asesores(self):
        ventana = self._crear_ventana("RECA - Asesores")
        AppEntidad(
            ventana,
            table="asesores",
            columns=("nombre", "email", "telefono", "sede", "gestor"),
            labels={
                "nombre": "Nombre",
                "email": "Email",
                "telefono": "Teléfono",
                "sede": "Sede",
                "gestor": "Gestor",
            },
            widths={
                "nombre": 220,
                "email": 220,
                "telefono": 120,
                "sede": 120,
                "gestor": 180,
            },
            form_config=[
                ("nombre", "Nombre *", True, "entry"),
                ("email", "Email", False, "entry"),
                ("telefono", "Teléfono", False, "entry"),
                ("sede", "Sede", False, "entry"),
                ("gestor", "Gestor", False, "entry"),
            ],
            key_field="nombre",
            title="RECA - Asesores",
            order_by="nombre",
        )

    def _abrir_gestores(self):
        ventana = self._crear_ventana("RECA - Gestores")
        AppEntidad(
            ventana,
            table="gestores",
            columns=("nombre", "email", "telefono", "sede", "localidades"),
            labels={
                "nombre": "Nombre",
                "email": "Email",
                "telefono": "Teléfono",
                "sede": "Sede",
                "localidades": "Localidades",
            },
            widths={
                "nombre": 220,
                "email": 220,
                "telefono": 120,
                "sede": 120,
                "localidades": 300,
            },
            form_config=[
                ("nombre", "Nombre *", True, "entry"),
                ("email", "Email", False, "entry"),
                ("telefono", "Teléfono", False, "entry"),
                ("sede", "Sede", False, "entry"),
                ("localidades", "Localidades", False, "text"),
            ],
            key_field="nombre",
            title="RECA - Gestores",
            order_by="nombre",
        )

    def _abrir_profesionales(self):
        ventana = self._crear_ventana("RECA - Profesionales")
        AppEntidad(
            ventana,
            table="profesionales",
            columns=("nombre_profesional", "correo_profesional", "programa", "antiguedad", "usuario_login"),
            labels={
                "nombre_profesional": "Nombre",
                "correo_profesional": "Email",
                "programa": "Programa",
                "antiguedad": "Antigüedad",
                "usuario_login": "Usuario Login",
            },
            widths={
                "nombre_profesional": 220,
                "correo_profesional": 220,
                "programa": 160,
                "antiguedad": 120,
                "usuario_login": 140,
            },
            form_config=[
                ("nombre_profesional", "Nombre *", True, "entry"),
                ("correo_profesional", "Email", False, "entry"),
                ("programa", "Programa", False, "entry"),
                ("antiguedad", "Antigüedad", False, "entry"),
                ("usuario_login", "Usuario Login", False, "entry"),
            ],
            key_field="id",
            title="RECA - Profesionales",
            order_by="nombre_profesional",
        )

    def _abrir_interpretes(self):
        ventana = self._crear_ventana("RECA - Interpretes")
        AppEntidad(
            ventana,
            table="interpretes",
            columns=("nombre",),
            labels={
                "nombre": "Nombre",
            },
            widths={
                "nombre": 300,
            },
            form_config=[
                ("nombre", "Nombre *", True, "entry"),
            ],
            key_field="nombre",
            title="RECA - Interpretes",
            order_by="nombre",
        )



# ============================================

# APLICACIÓN PRINCIPAL

# ============================================



class AppRECA:

    """Aplicación principal de gestión de empresas"""



    # Configuración de columnas de la tabla

    COLUMNAS = (

        "nombre_empresa", "nit_empresa", "ciudad_empresa", "estado", "zona_empresa",

        "profesional_asignado", "asesor", "contacto_empresa", "telefono_empresa", "sede_empresa"

    )



    ANCHOS_COLUMNAS = {

        "nombre_empresa": 250,

        "nit_empresa": 100,

        "ciudad_empresa": 100,

        "estado": 100,

        "zona_empresa": 120,

        "profesional_asignado": 150,

        "asesor": 130,

        "contacto_empresa": 150,

        "telefono_empresa": 100,

        "sede_empresa": 100

    }

    COLUMN_LABELS = {
        "nombre_empresa": "Nombre Empresa",
        "nit_empresa": "NIT",
        "ciudad_empresa": "Ciudad",
        "estado": "Estado",
        "zona_empresa": "Zona Compensar",
        "profesional_asignado": "Profesional Asignado",
        "asesor": "Asesor",
        "contacto_empresa": "Contacto(s)",
        "telefono_empresa": "Teléfono(s)",
        "sede_empresa": "Sede Empresa",
    }



    # Tamaño del lote para paginación (máximo de Supabase es 1000)

    BATCH_SIZE = 1000



    def __init__(self, root, progress_callback=None, on_ready=None):
        """
        Inicializa la aplicacion

        Args:
            root: Ventana raiz de tkinter
            progress_callback: Funcion para reportar avance
            on_ready: Callback cuando la app termina de iniciar
        """
        self.root = root
        self.root.title("RECA - Gestion de Empresas")
        self.root.geometry("1400x750")
        _maximize_window(self.root)

        if os.path.exists(LOGO_PATH):
            try:
                icon = tk.PhotoImage(file=LOGO_PATH)
                self.root.iconphoto(True, icon)
                self._window_icon = icon
            except Exception:
                self._window_icon = None

        self._progress_callback = progress_callback
        self._on_ready = on_ready
        self._progress_value = 0
        self._report_progress("Iniciando interfaz...", 5)

        # Inicializar variables
        self._report_progress("Conectando a Supabase...", 15)
        self.supabase = conectar_supabase()
        self.empresas_actuales = []
        self._empresa_por_id = {}
        self.empresa_seleccionada = None
        self._offset = 0
        self._all_loaded = False
        self._is_loading = False
        self._search_term = ""
        self._search_field = "Todos"
        self._sort_state = {}
        self._profesionales_cache = []
        self._filters = {
            "profesional_asignado": None,
            "asesor": None,
            "caja_compensacion": None,
            "zona_empresa": None,
            "estado": None,
        }
        self._filtros_visible = False
        self._filtros_loaded = False
        self._filtros_loading = False
        self._latest_version = None

        # Crear interfaz y cargar datos
        self._report_progress("Construyendo interfaz...", 35)
        self.crear_interfaz()
        self._report_progress("Cargando empresas...", 55)
        self.cargar_todas_empresas()

    def _report_progress(self, message, value=None):
        if value is not None:
            self._progress_value = value
        if self._progress_callback:
            self._progress_callback(message, self._progress_value)

    def _finish_ready(self):
        if self._on_ready:
            self._report_progress("Listo", 100)
            callback = self._on_ready
            self._on_ready = None
            callback()

    def crear_interfaz(self):

        """Construye la interfaz principal de la aplicación"""



        # Header

        self._crear_header()



        # Barra de búsqueda

        self._crear_barra_busqueda()

        # Filtros avanzados

        self._crear_filtros()



        # Tabla de resultados

        self._crear_tabla()



        # Botones de acción

        self._crear_botones_accion()

        # Footer

        self._crear_footer()

        # Cargar version publicada (en background)
        self._fetch_latest_version_async()

        # Atajos de teclado
        self.root.bind("<Control-n>", lambda e: self.nueva_empresa())
        self.root.bind("<Control-f>", lambda e: self.search_entry.focus_set())
        self.root.bind("<F5>", lambda e: self.cargar_todas_empresas())


    def _crear_header(self):
        """Crea el encabezado de la aplicacion"""
        header = tk.Frame(self.root, bg=COLOR_PURPLE, height=72)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        header.grid_rowconfigure(0, weight=1)
        header.grid_columnconfigure(1, weight=1)

        self.header_logo = None
        if os.path.exists(LOGO_PATH):
            try:
                logo = tk.PhotoImage(file=LOGO_PATH)
                max_size = 48
                scale = max(1, logo.width() // max_size, logo.height() // max_size)
                if scale > 1:
                    logo = logo.subsample(scale, scale)
                self.header_logo = logo
                tk.Label(header, image=self.header_logo, bg=COLOR_PURPLE).grid(
                    row=0, column=0, padx=(SP_MD, SP_SM), pady=SP_SM, sticky="w"
                )
            except Exception:
                self.header_logo = None

        tk.Label(
            header,
            text="RECA - Gestion de Empresas",
            font=FONT_H1,
            bg=COLOR_PURPLE,
            fg="white",
        ).grid(row=0, column=1, sticky="w", padx=(0, 16))

    def _crear_barra_busqueda(self):
        """Crea la barra de busqueda"""
        search_frame = tk.Frame(self.root, bg=COLOR_LIGHT_BG, height=70)
        search_frame.pack(fill=tk.X, pady=10, padx=20)

        # Campo de busqueda
        tk.Label(
            search_frame,
            text="Buscar:",
            font=FONT_BODY,
            bg=COLOR_LIGHT_BG,
            fg=COLOR_PURPLE,
        ).pack(side=tk.LEFT, padx=5)

        self.search_var = tk.StringVar()
        self.search_entry = ttk.Combobox(
            search_frame,
            textvariable=self.search_var,
            values=[],
            font=FONT_BODY,
            width=38,
            state="normal",
        )
        self.search_entry.pack(side=tk.LEFT, padx=SP_XS)
        self.search_entry.bind("<Return>", lambda e: self.buscar_empresas())
        self.search_entry.bind("<KeyRelease>", lambda e: self._update_autocomplete())

        # Placeholder
        self._placeholder = "Escribe para buscar..."
        self._placeholder_active = True
        self.search_entry.set(self._placeholder)
        self.search_entry.config(foreground="#999999")

        def _on_focus_in(_e):
            if self._placeholder_active:
                self.search_entry.set("")
                self.search_entry.config(foreground=COLOR_TEXT)
                self._placeholder_active = False

        def _on_focus_out(_e):
            if not self.search_var.get().strip():
                self.search_entry.set(self._placeholder)
                self.search_entry.config(foreground="#999999")
                self._placeholder_active = True

        self.search_entry.bind("<FocusIn>", _on_focus_in, add="+")
        self.search_entry.bind("<FocusOut>", _on_focus_out, add="+")

        # Selector de campo
        tk.Label(
            search_frame,
            text="En:",
            font=FONT_BODY,
            bg=COLOR_LIGHT_BG,
            fg=COLOR_PURPLE,
        ).pack(side=tk.LEFT, padx=5)

        self.campo_busqueda = ttk.Combobox(
            search_frame,
            values=["Todos", "Nombre", "NIT", "Ciudad", "Profesional"],
            state="readonly",
            width=12,
        )
        self.campo_busqueda.set("Todos")
        self.campo_busqueda.pack(side=tk.LEFT, padx=5)
        self.campo_busqueda.bind("<<ComboboxSelected>>", lambda e: self._on_search_field_change())

        # Botones de busqueda
        _make_button(search_frame, "Buscar", self.buscar_empresas, style="secondary", font=FONT_BODY_BOLD).pack(side=tk.LEFT, padx=SP_XS)
        _make_button(search_frame, "Limpiar", self.limpiar_busqueda, style="outline", font=FONT_BODY_BOLD).pack(side=tk.LEFT, padx=SP_XS)
        _make_button(search_frame, "Filtros", self._toggle_filtros, style="neutral", font=FONT_BODY_BOLD).pack(side=tk.LEFT, padx=SP_XS)

    def _crear_footer(self):
        footer = tk.Frame(self.root, bg=COLOR_LIGHT_BG, height=24)
        footer.pack(fill=tk.X, side=tk.BOTTOM)
        footer.pack_propagate(False)
        self.version_label = tk.Label(
            footer,
            text=f"Version: {APP_VERSION} | GitHub: --",
            font=FONT_SMALL,
            bg=COLOR_LIGHT_BG,
            fg="#666666",
        )
        self.version_label.pack(side=tk.LEFT, padx=8)

    def _fetch_latest_version_async(self):
        def worker():
            latest = None
            try:
                release = _get_latest_release()
                if release:
                    latest = release.get("tag_name", "") or None
            except Exception:
                LOG.exception("Error fetching latest release for footer")
            self.root.after(0, lambda: self._update_version_label(latest))

        threading.Thread(target=worker, daemon=True).start()

    def _update_version_label(self, latest):
        self._latest_version = latest
        latest_text = latest if latest else "--"
        if self.version_label.winfo_exists():
            self.version_label.config(text=f"Version: {APP_VERSION} | GitHub: {latest_text}")

    def _crear_filtros(self):
        self.filtros_frame = tk.Frame(self.root, bg=COLOR_LIGHT_BG)

        tk.Label(
            self.filtros_frame,
            text="Filtros:",
            font=FONT_H3,
            bg=COLOR_LIGHT_BG,
            fg=COLOR_PURPLE,
        ).grid(row=0, column=0, sticky="w", padx=5, pady=5)

        self.filtro_profesional = self._crear_filtro_combo(
            self.filtros_frame, "Profesional", 0, 1
        )
        self.filtro_asesor = self._crear_filtro_combo(
            self.filtros_frame, "Asesor", 0, 3
        )
        self.filtro_caja = self._crear_filtro_combo(
            self.filtros_frame, "Caja Compensación", 0, 5
        )
        self.filtro_zona = self._crear_filtro_combo(
            self.filtros_frame, "Zona Compensar", 1, 1
        )
        self.filtro_estado = self._crear_filtro_combo(
            self.filtros_frame, "Estado", 1, 3
        )

        btn_aplicar = _make_button(self.filtros_frame, "Aplicar", self.aplicar_filtros, style="secondary", font=FONT_BODY_BOLD)
        btn_aplicar.grid(row=1, column=5, padx=SP_XS, pady=SP_XS, sticky="w")

        btn_limpiar = _make_button(self.filtros_frame, "Limpiar Filtros", self.limpiar_filtros, style="outline", font=FONT_BODY_BOLD)
        btn_limpiar.grid(row=1, column=6, padx=SP_XS, pady=SP_XS, sticky="w")

        self.filtros_frame.grid_columnconfigure(2, minsize=8)
        self.filtros_frame.grid_columnconfigure(4, minsize=8)

    def _crear_filtro_combo(self, parent, label, row, col):
        tk.Label(
            parent,
            text=f"{label}:",
            font=FONT_BODY,
            bg=COLOR_LIGHT_BG,
            fg=COLOR_PURPLE,
        ).grid(row=row, column=col, sticky="e", padx=5, pady=5)

        combo = ttk.Combobox(
            parent,
            values=["Todos"],
            state="readonly",
            width=24,
            font=FONT_BODY,
        )
        combo.set("Todos")
        combo.grid(row=row, column=col + 1, sticky="w", padx=5, pady=5)
        return combo

    def _crear_tabla(self):
        """Crea la tabla de empresas con scrollbars"""
        tabla_frame = tk.Frame(self.root)
        tabla_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Contador de resultados
        self.contador_label = tk.Label(
            tabla_frame,
            text="Resultados: 0 empresas",
            font=FONT_BODY_BOLD
        )
        self.contador_label.grid(row=0, column=0, sticky="w", pady=5)

        # Treeview
        self.tree = ttk.Treeview(
            tabla_frame,
            columns=self.COLUMNAS,
            show="headings",
            height=18
        )

        # Configurar columnas
        for col in self.COLUMNAS:
            # Encabezado con funcion de ordenamiento
            self.tree.heading(
                col,
                text=self.COLUMN_LABELS.get(col, col.replace("_", " ").title()),
                command=lambda c=col: self.ordenar_columna(c)
            )
            # Ancho de columna
            self.tree.column(col, width=self.ANCHOS_COLUMNAS.get(col, 100))

        def on_scrollbar(*args):
            self.tree.yview(*args)
            self._maybe_load_next()

        # Scrollbars
        scroll_y = ttk.Scrollbar(tabla_frame, orient=tk.VERTICAL, command=on_scrollbar)
        scroll_x = ttk.Scrollbar(tabla_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscroll=scroll_y.set, xscroll=scroll_x.set)

        # Grid layout
        self.tree.grid(row=1, column=0, sticky="nsew")
        scroll_y.grid(row=1, column=1, sticky="ns")
        scroll_x.grid(row=2, column=0, sticky="ew")

        tabla_frame.grid_rowconfigure(1, weight=1)
        tabla_frame.grid_columnconfigure(0, weight=1)

        # Eventos
        self.tree.bind("<ButtonRelease-1>", self.seleccionar)
        self.tree.bind("<Double-1>", self.abrir_editar)
        self.tree.bind("<MouseWheel>", lambda e: self._maybe_load_next())
        self.tree.bind("<Button-4>", lambda e: self._maybe_load_next())
        self.tree.bind("<Button-5>", lambda e: self._maybe_load_next())
        self.tree.bind("<KeyRelease-Next>", lambda e: self._maybe_load_next())
        self.tree.bind("<Configure>", lambda e: self._maybe_load_next())

        # Filas alternadas
        self.tree.tag_configure("oddrow", background=TREE_ROW_ALT)
        self.tree.tag_configure("evenrow", background=COLOR_WHITE)

    def _reset_paginacion(self):
        self._offset = 0
        self._all_loaded = False
        self._is_loading = False
        self.empresas_actuales = []
        self._empresa_por_id = {}
        self.empresa_seleccionada = None
        self._limpiar_tabla()
        self._update_contador()
        self._update_autocomplete_values()

    def _build_or_filter(self, term):
        safe_term = self._sanitize_search_term(term)
        columnas = [
            "nombre_empresa",
            "nit_empresa",
            "ciudad_empresa",
            "estado",
            "zona_empresa",
            "profesional_asignado",
            "asesor",
            "contacto_empresa",
            "telefono_empresa",
            "sede_empresa",
        ]
        return ",".join([f"{col}.ilike.%{safe_term}%" for col in columnas])

    def _sanitize_search_term(self, term):
        safe_term = re.sub(r"[^\w\s@.-]", " ", term, flags=re.UNICODE)
        safe_term = re.sub(r"\s+", " ", safe_term).strip()
        return safe_term

    def _load_next_page(self):
        if not self.supabase or self._all_loaded or self._is_loading:
            return
        self._is_loading = True
        try:
            query = self.supabase.table("empresas").select("*").order("id", desc=False)
            for col, value in self._filters.items():
                if value and value != "Todos":
                    query = query.eq(col, value)
            if self._search_term:
                if self._search_field == "Todos":
                    query = query.or_(self._build_or_filter(self._search_term))
                else:
                    campo_columna = {
                        "Nombre": "nombre_empresa",
                        "NIT": "nit_empresa",
                        "Ciudad": "ciudad_empresa",
                        "Profesional": "profesional_asignado",
                    }.get(self._search_field, "nombre_empresa")
                    query = query.ilike(campo_columna, f"%{self._search_term}%")
            query = query.range(self._offset, self._offset + self.BATCH_SIZE - 1)
            response = query.execute()
            data = response.data or []
            if not data:
                self._all_loaded = True
                return
            self._offset += len(data)
            self.empresas_actuales.extend(data)
            self._mostrar_empresas(data)
            next_value = min(90, max(self._progress_value, 60) + 10)
            self._report_progress(f"Cargando empresas... ({len(self.empresas_actuales)})", next_value)
            if len(data) < self.BATCH_SIZE:
                self._all_loaded = True
            self._update_contador()
            self._update_autocomplete_values()
            if not self._all_loaded:
                self.root.after(0, self._maybe_load_next)
        except Exception as e:
            LOG.exception("Error cargando empresas")
            messagebox.showerror("Error", f"Error cargando empresas: {e}")
            self._all_loaded = True
        finally:
            self._is_loading = False

    def _maybe_load_next(self):
        if self._all_loaded or self._is_loading:
            return
        first, last = self.tree.yview()
        if last >= 0.98:
            self._load_next_page()

    def _update_contador(self):
        self.contador_label.config(text=f"Resultados: {len(self.empresas_actuales)} empresas")

    def _crear_botones_accion(self):
        """Crea los botones de accion principales"""
        btn_frame = tk.Frame(self.root, bg=COLOR_LIGHT_BG)
        btn_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=SP_LG, pady=SP_SM)

        _make_button(btn_frame, "Nueva Empresa", self.nueva_empresa, style="primary").pack(side=tk.LEFT, padx=SP_XS)
        _make_button(btn_frame, "Importar Excel", self.importar_empresas_excel, style="info").pack(side=tk.LEFT, padx=SP_XS)
        _make_button(btn_frame, "Editar", self.editar_empresa, style="secondary").pack(side=tk.LEFT, padx=SP_XS)
        _make_button(btn_frame, "Refrescar", self.cargar_todas_empresas, style="outline").pack(side=tk.LEFT, padx=SP_XS)
        _make_button(btn_frame, "Eliminar", self.eliminar_empresa, style="danger").pack(side=tk.RIGHT, padx=SP_XS)

    def _build_excel_index_map(self, headers):
        header_keys = [_header_key(h) for h in headers]

        def pick(keys, default=None):
            for idx, key in enumerate(header_keys):
                if key in keys:
                    return idx
            return default

        cargo_indices = [idx for idx, key in enumerate(header_keys) if key == "cargo"]
        cargo_contacto_idx = pick({"cargocontacto"}, None)
        cargo_responsable_idx = pick({"cargoresponsable"}, None)

        if cargo_contacto_idx is None:
            cargo_contacto_idx = cargo_indices[0] if cargo_indices else 6
        if cargo_responsable_idx is None:
            cargo_responsable_idx = cargo_indices[1] if len(cargo_indices) > 1 else cargo_contacto_idx

        return {
            "nombre_empresa": pick({"nombre", "nombreempresa"}, 0),
            "nit_empresa": pick({"nit", "nitempresa"}, 1),
            "direccion_empresa": pick({"direccion", "direccionempresa"}, 2),
            "ciudad_empresa": pick({"ciudad", "ciudadempresa"}, 3),
            "correo_1": pick({"correo1"}, 4),
            "contacto_empresa": pick({"contacto", "contactoempresa"}, 5),
            "cargo_contacto": cargo_contacto_idx,
            "sede_empresa": pick({"sede", "sedeempresa"}, 7),
            "telefono_empresa": pick({"telefono", "telefonoempresa"}, 8),
            "responsable_visita": pick({"responsabledelavisita", "responsablevisita"}, 9),
            "cargo_responsable": cargo_responsable_idx,
            "asesor": pick({"asesor"}, 11),
            "correo_asesor": pick({"correodeasesor", "correoasesor"}, 12),
            "zona_empresa": pick({"zona", "zonaempresa"}, 13),
            "caja_compensacion": pick({"cajadecompensacion", "cajacompensacion"}, 14),
            "profesional_asignado": pick({"profesionalasignado"}, 15),
            "correo_profesional": pick({"correoprofesional", "correoprofesionalasignado"}, 16),
            "estado": pick({"estado"}, 17),
            "observaciones": pick({"observaciones", "observacion"}, 18),
        }

    def _map_excel_row(self, row, idx):
        def cell(index):
            if index is None:
                return None
            if index < 0 or index >= len(row):
                return None
            return row[index]

        cargo = _clean_text(cell(idx.get("cargo_responsable"))) or _clean_text(cell(idx.get("cargo_contacto")))

        return {
            "nombre_empresa": _clean_text(cell(idx.get("nombre_empresa"))),
            "nit_empresa": _clean_text(cell(idx.get("nit_empresa"))),
            "direccion_empresa": _clean_text(cell(idx.get("direccion_empresa"))),
            "ciudad_empresa": _clean_text(cell(idx.get("ciudad_empresa"))),
            "correo_1": _clean_text(cell(idx.get("correo_1"))),
            "contacto_empresa": _clean_text(cell(idx.get("contacto_empresa"))),
            "cargo": cargo,
            "sede_empresa": _clean_text(cell(idx.get("sede_empresa"))),
            "telefono_empresa": _clean_text(cell(idx.get("telefono_empresa"))),
            "responsable_visita": _clean_text(cell(idx.get("responsable_visita"))),
            "asesor": _clean_text(cell(idx.get("asesor"))),
            "correo_asesor": _clean_text(cell(idx.get("correo_asesor"))),
            "zona_empresa": _clean_text(cell(idx.get("zona_empresa"))),
            "caja_compensacion": _clean_text(cell(idx.get("caja_compensacion"))),
            "profesional_asignado": _clean_text(cell(idx.get("profesional_asignado"))),
            "correo_profesional": _clean_text(cell(idx.get("correo_profesional"))),
            "estado": _clean_text(cell(idx.get("estado"))),
            "observaciones": _clean_text(cell(idx.get("observaciones"))),
        }

    def _fetch_existing_empresa_pairs(self):
        existing_pairs = set()
        offset = 0
        while True:
            query = (
                self.supabase.table("empresas")
                .select("nit_empresa,nombre_empresa")
                .order("id", desc=False)
                .range(offset, offset + self.BATCH_SIZE - 1)
            )
            data = query.execute().data or []
            if not data:
                break
            for row in data:
                key_pair = (
                    _normalize_nit(row.get("nit_empresa")),
                    _normalize_name(row.get("nombre_empresa")),
                )
                existing_pairs.add(key_pair)
            offset += len(data)
            if len(data) < self.BATCH_SIZE:
                break
        return existing_pairs

    def _crear_tab_resumen_importacion(self, parent, title, rows, with_checkbox=False, on_selection_change=None):
        frame = tk.Frame(parent)
        parent.add(frame, text=f"{title} ({len(rows)})")

        selection_state = None
        if with_checkbox:
            columns = ("seleccion", "nombre_empresa", "nit_empresa", "ciudad_empresa", "estado")
        else:
            columns = ("nombre_empresa", "nit_empresa", "ciudad_empresa", "estado")

        tree = ttk.Treeview(frame, columns=columns, show="headings", height=12)
        if with_checkbox:
            tree.heading("seleccion", text="Subir")
            tree.column("seleccion", width=70, anchor="center")
        tree.heading("nombre_empresa", text="Nombre Empresa")
        tree.heading("nit_empresa", text="NIT")
        tree.heading("ciudad_empresa", text="Ciudad")
        tree.heading("estado", text="Estado")
        tree.column("nombre_empresa", width=420)
        tree.column("nit_empresa", width=160)
        tree.column("ciudad_empresa", width=160)
        tree.column("estado", width=120)

        scroll_y = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scroll_y.set)
        tree.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        max_preview = len(rows) if with_checkbox else 500
        if with_checkbox:
            selection_state = {"selected": {}, "rows": {}}

        for row in rows[:max_preview]:
            if with_checkbox:
                item_id = tree.insert(
                    "",
                    tk.END,
                    values=(
                        "[x]",
                        row.get("nombre_empresa", ""),
                        row.get("nit_empresa", ""),
                        row.get("ciudad_empresa", ""),
                        row.get("estado", ""),
                    ),
                )
                selection_state["selected"][item_id] = True
                selection_state["rows"][item_id] = row
            else:
                tree.insert(
                    "",
                    tk.END,
                    values=(
                        row.get("nombre_empresa", ""),
                        row.get("nit_empresa", ""),
                        row.get("ciudad_empresa", ""),
                        row.get("estado", ""),
                    ),
                )

        if len(rows) > max_preview:
            tk.Label(
                frame,
                text=f"Mostrando {max_preview} de {len(rows)} registros",
                font=FONT_SMALL,
                fg="#666666",
            ).grid(row=1, column=0, sticky="w", pady=(6, 0))

        if with_checkbox:
            def _notify():
                if on_selection_change:
                    on_selection_change(selection_state)
                dynamic_callback = selection_state.get("on_change")
                if dynamic_callback:
                    dynamic_callback(selection_state)

            def toggle_item(item_id):
                if not item_id:
                    return
                current = selection_state["selected"].get(item_id, True)
                selection_state["selected"][item_id] = not current
                values = list(tree.item(item_id, "values"))
                if values:
                    values[0] = "[x]" if not current else "[ ]"
                    tree.item(item_id, values=values)
                _notify()

            def on_click(event):
                if tree.identify("region", event.x, event.y) != "cell":
                    return
                if tree.identify_column(event.x) != "#1":
                    return
                toggle_item(tree.identify_row(event.y))

            tree.bind("<Button-1>", on_click)

            controls = tk.Frame(frame)
            controls.grid(row=2, column=0, sticky="w", pady=(8, 0))

            def set_all(value):
                for item_id in tree.get_children(""):
                    selection_state["selected"][item_id] = value
                    values = list(tree.item(item_id, "values"))
                    if values:
                        values[0] = "[x]" if value else "[ ]"
                        tree.item(item_id, values=values)
                _notify()

            _make_button(controls, "Marcar todos", lambda: set_all(True), style="primary", font=FONT_SMALL).pack(side=tk.LEFT, padx=(0, SP_SM))
            _make_button(controls, "Desmarcar todos", lambda: set_all(False), style="outline", font=FONT_SMALL).pack(side=tk.LEFT)

            _notify()

        return selection_state

    def _confirmar_subida_importacion_excel(self, window, empresas_a_subir):
        if not empresas_a_subir:
            messagebox.showinfo("Importacion", "No hay empresas nuevas para subir.")
            return

        confirmar = messagebox.askyesno(
            "Confirmar subida",
            f"Se subiran {len(empresas_a_subir)} empresas nuevas a la nube.\n\nDeseas continuar?",
        )
        if not confirmar:
            return

        try:
            inserted = 0
            chunk_size = 200
            for idx in range(0, len(empresas_a_subir), chunk_size):
                chunk = empresas_a_subir[idx:idx + chunk_size]
                self.supabase.table("empresas").insert(chunk).execute()
                inserted += len(chunk)

            LOG.info("Importacion Excel empresas completada. Insertadas: %s", inserted)
            messagebox.showinfo("Importacion", f"Importacion completada.\nEmpresas subidas: {inserted}")
            if window and window.winfo_exists():
                window.destroy()
            self.cargar_todas_empresas()
        except Exception as e:
            LOG.exception("Error subiendo importacion Excel de empresas")
            messagebox.showerror("Error", f"No se pudo completar la importacion:\n{e}")

    def _mostrar_resumen_importacion_excel(self, file_path, nuevas, repetidas_bd, repetidas_archivo, filas_validas):
        window = tk.Toplevel(self.root)
        window.title("Resumen importacion Excel")
        window.geometry("1200x760")
        _maximize_window(window)

        wrap = tk.Frame(window, bg=COLOR_LIGHT_BG)
        wrap.pack(fill=tk.BOTH, expand=True, padx=14, pady=12)

        tk.Label(
            wrap,
            text="Resumen de importacion de empresas",
            font=FONT_H2,
            bg=COLOR_LIGHT_BG,
            fg=COLOR_PURPLE,
        ).pack(anchor="w")

        tk.Label(
            wrap,
            text=f"Archivo: {file_path}",
            font=FONT_BODY,
            bg=COLOR_LIGHT_BG,
            fg="#4a4a4a",
            wraplength=1100,
            justify="left",
        ).pack(anchor="w", pady=(2, 8))

        summary = (
            f"Filas validas: {filas_validas}    "
            f"Nuevas: {len(nuevas)}    "
            f"Repetidas en nube (NIT + Nombre): {len(repetidas_bd)}    "
            f"Repetidas dentro del archivo: {len(repetidas_archivo)}"
        )
        tk.Label(
            wrap,
            text=summary,
            font=FONT_BODY_BOLD,
            bg=COLOR_LIGHT_BG,
            fg=COLOR_TEAL,
        ).pack(anchor="w", pady=(0, 10))

        notebook = ttk.Notebook(wrap)
        notebook.pack(fill=tk.BOTH, expand=True)
        nuevas_selection = self._crear_tab_resumen_importacion(
            notebook,
            "Nuevas",
            nuevas,
            with_checkbox=True,
        )
        self._crear_tab_resumen_importacion(notebook, "Repetidas en nube", repetidas_bd)
        self._crear_tab_resumen_importacion(notebook, "Repetidas en archivo", repetidas_archivo)

        btn_row = tk.Frame(wrap, bg=COLOR_LIGHT_BG)
        btn_row.pack(fill=tk.X, pady=(10, 0))

        def selected_rows():
            if not nuevas_selection:
                return list(nuevas)
            return [
                row
                for item_id, row in nuevas_selection["rows"].items()
                if nuevas_selection["selected"].get(item_id, False)
            ]

        def refresh_confirm_button(_=None):
            count = len(selected_rows())
            confirm_btn.config(
                text=f"Confirmar y subir {count}",
                state=tk.NORMAL if count else tk.DISABLED,
            )

        if nuevas_selection is not None:
            nuevas_selection["on_change"] = refresh_confirm_button

        confirm_btn = _make_button(
            btn_row,
            f"Confirmar y subir {len(nuevas)}",
            lambda: self._confirmar_subida_importacion_excel(window, selected_rows()),
            style="primary",
            font=FONT_BODY_BOLD,
            state=tk.NORMAL if nuevas else tk.DISABLED,
        )
        confirm_btn.pack(side=tk.LEFT, padx=(0, SP_SM))
        refresh_confirm_button()

        _make_button(btn_row, "Cerrar", window.destroy, style="secondary", font=FONT_BODY_BOLD).pack(side=tk.LEFT)

    def importar_empresas_excel(self):
        if not self.supabase:
            return

        file_path = filedialog.askopenfilename(
            title="Selecciona archivo Excel de empresas",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
        )
        if not file_path:
            return

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb[wb.sheetnames[0]]
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                messagebox.showwarning("Importacion", "El archivo no tiene encabezados.")
                return

            idx = self._build_excel_index_map(list(header_row))
            existing_pairs = self._fetch_existing_empresa_pairs()

            nuevas = []
            repetidas_bd = []
            repetidas_archivo = []
            seen_in_file = set()
            filas_validas = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                record = self._map_excel_row(row, idx)
                if not any(record.values()):
                    continue

                key_pair = (
                    _normalize_nit(record.get("nit_empresa")),
                    _normalize_name(record.get("nombre_empresa")),
                )
                if not key_pair[0] and not key_pair[1]:
                    continue
                filas_validas += 1

                if key_pair in seen_in_file:
                    repetidas_archivo.append(record)
                    continue

                seen_in_file.add(key_pair)
                if key_pair in existing_pairs:
                    repetidas_bd.append(record)
                else:
                    nuevas.append(record)

            self._mostrar_resumen_importacion_excel(
                file_path=file_path,
                nuevas=nuevas,
                repetidas_bd=repetidas_bd,
                repetidas_archivo=repetidas_archivo,
                filas_validas=filas_validas,
            )
        except Exception as e:
            LOG.exception("Error importando archivo Excel de empresas")
            messagebox.showerror(
                "Error",
                "No se pudo procesar el archivo Excel.\n"
                f"Detalle: {e}",
            )

    def cargar_todas_empresas(self):
        """
        Carga empresas con paginacion por lotes.
        """
        if not self.supabase:
            self._finish_ready()
            return

        self._search_term = ""
        self._search_field = "Todos"
        self._reset_paginacion()
        self._load_next_page()

        if self._all_loaded and not self.empresas_actuales:
            messagebox.showinfo("Info", "No hay empresas en la base de datos")
        self._finish_ready()


    def buscar_empresas(self):
        """Busca empresas segun el termino y campo seleccionado"""
        if not self.supabase:
            return

        termino = self.search_entry.get().strip()
        if not termino or self._placeholder_active:
            self.cargar_todas_empresas()
            return
        termino = self._sanitize_search_term(termino)
        if not termino:
            messagebox.showwarning("Aviso", "Ingresa un texto de busqueda valido")
            return

        self._search_term = termino
        self._search_field = self.campo_busqueda.get()
        self._reset_paginacion()
        self._load_next_page()

        if self._all_loaded and not self.empresas_actuales:
            messagebox.showinfo("Info", "No hay resultados")

    def limpiar_busqueda(self):
        """Limpia el campo de busqueda y recarga todas las empresas"""
        self.search_entry.set(self._placeholder)
        self.search_entry.config(foreground="#999999")
        self._placeholder_active = True
        self.campo_busqueda.set("Todos")
        self.cargar_todas_empresas()
        self._update_autocomplete()

    def _on_search_field_change(self):
        self._search_field = self.campo_busqueda.get()
        self._update_autocomplete_values()
        self._update_autocomplete()

    def _update_autocomplete_values(self):
        if self._search_field == "Profesional":
            profesionales = []
            for empresa in self.empresas_actuales:
                valor = (empresa.get("profesional_asignado") or "").strip()
                if valor:
                    profesionales.append(valor)
            self._profesionales_cache = sorted(set(profesionales))
        else:
            self._profesionales_cache = []
        if self._search_field != "Profesional":
            self.search_entry["values"] = []

    def _update_autocomplete(self):
        if self.campo_busqueda.get() != "Profesional":
            return
        term = self.search_entry.get().strip().lower()
        if not term:
            self.search_entry["values"] = self._profesionales_cache
            return
        matches = [p for p in self._profesionales_cache if term in p.lower()]
        self.search_entry["values"] = matches

    def _toggle_filtros(self):
        if self._filtros_visible:
            self.filtros_frame.pack_forget()
            self._filtros_visible = False
            return
        self.filtros_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        self._filtros_visible = True
        if not self._filtros_loaded:
            self._load_filter_options()

    def _load_filter_options(self):
        if not self.supabase or self._filtros_loading:
            return
        self._filtros_loading = True
        options = {
            "profesional_asignado": set(),
            "asesor": set(),
            "caja_compensacion": set(),
            "zona_empresa": set(),
            "estado": set(),
        }
        offset = 0
        try:
            while True:
                query = (self.supabase.table("empresas")
                         .select("profesional_asignado, asesor, caja_compensacion, zona_empresa, estado")
                         .order("id", desc=False)
                         .range(offset, offset + self.BATCH_SIZE - 1))
                data = query.execute().data or []
                if not data:
                    break
                for row in data:
                    for key in options:
                        valor = (row.get(key) or "").strip()
                        if valor:
                            options[key].add(valor)
                offset += len(data)
                if len(data) < self.BATCH_SIZE:
                    break
        except Exception:
            LOG.exception("Error cargando opciones de filtros")
        finally:
            self._filtros_loading = False

        self.filtro_profesional["values"] = ["Todos"] + sorted(options["profesional_asignado"])
        self.filtro_asesor["values"] = ["Todos"] + sorted(options["asesor"])
        self.filtro_caja["values"] = ["Todos"] + sorted(options["caja_compensacion"])
        self.filtro_zona["values"] = ["Todos"] + sorted(options["zona_empresa"])
        self.filtro_estado["values"] = ["Todos"] + sorted(options["estado"])
        self._filtros_loaded = True

    def aplicar_filtros(self):
        self._filters = {
            "profesional_asignado": self.filtro_profesional.get(),
            "asesor": self.filtro_asesor.get(),
            "caja_compensacion": self.filtro_caja.get(),
            "zona_empresa": self.filtro_zona.get(),
            "estado": self.filtro_estado.get(),
        }
        self._reset_paginacion()
        self._load_next_page()

    def limpiar_filtros(self):
        self.filtro_profesional.set("Todos")
        self.filtro_asesor.set("Todos")
        self.filtro_caja.set("Todos")
        self.filtro_zona.set("Todos")
        self.filtro_estado.set("Todos")
        for key in self._filters:
            self._filters[key] = None
        self._reset_paginacion()
        self._load_next_page()

    def _sort_key(self, value):
        if value is None:
            return (1, "")
        value = str(value).strip()
        if not value:
            return (1, "")
        try:
            return (0, float(value.replace(",", ".")))
        except ValueError:
            return (1, value.lower())

    def ordenar_columna(self, col):
        """Ordena la tabla por la columna seleccionada (toggle asc/desc)"""
        ascending = self._sort_state.get(col, True)
        items = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]
        items.sort(key=lambda item: self._sort_key(item[0]), reverse=not ascending)
        for index, (_, k) in enumerate(items):
            self.tree.move(k, "", index)
        self._sort_state[col] = not ascending

    def seleccionar(self, event):
        """Maneja la seleccion de una empresa con un click"""
        selection = self.tree.selection()
        if selection:
            item = self.tree.item(selection[0])
            empresa_id = item["tags"][0] if item["tags"] else None
            if empresa_id is not None:
                empresa_id = str(empresa_id)
            self.empresa_seleccionada = self._empresa_por_id.get(empresa_id)

    def abrir_editar(self, event=None):

        """

        Abre el formulario de edición (doble click)



        Args:

            event: Evento de tkinter (opcional)

        """

        if not self.empresa_seleccionada:

            messagebox.showwarning("Aviso", "Selecciona una empresa primero")

            return



        ventana = FormularioEmpresa(self.root, self.supabase, self.empresa_seleccionada)

        self.root.wait_window(ventana)



        # Recargar si hubo cambios

        if ventana.resultado:

            self.cargar_todas_empresas()



    def nueva_empresa(self):

        """Abre el formulario para crear una nueva empresa"""

        if not self.supabase:

            return



        ventana = FormularioEmpresa(self.root, self.supabase)

        self.root.wait_window(ventana)



        # Recargar si se creó

        if ventana.resultado:

            self.cargar_todas_empresas()



    def editar_empresa(self):

        """Edita la empresa seleccionada (botón editar)"""

        if not self.empresa_seleccionada:

            messagebox.showwarning("Aviso", "Selecciona una empresa primero")

            return



        self.abrir_editar()



    def eliminar_empresa(self):

        """Elimina la empresa seleccionada (botón eliminar)"""

        if not self.empresa_seleccionada:

            messagebox.showwarning("Aviso", "Selecciona una empresa primero")

            return



        confirmar = messagebox.askyesno(

            "Confirmar",

            f"¿Eliminar '{self.empresa_seleccionada['nombre_empresa']}'?"

        )



        if confirmar:
            try:
                (self.supabase.table("empresas")
                    .delete()
                    .eq("id", self.empresa_seleccionada["id"])
                    .execute())
                LOG.info("Empresa eliminada (principal): %s", self.empresa_seleccionada.get("nombre_empresa"))
                messagebox.showinfo("?xito", "Empresa eliminada")
                self.cargar_todas_empresas()
            except Exception as e:
                LOG.exception("Error eliminando empresa (principal)")
                messagebox.showerror("Error", f"Error eliminando: {e}")



    def _limpiar_tabla(self):

        """Limpia todos los elementos de la tabla"""

        for item in self.tree.get_children():

            self.tree.delete(item)



    def _mostrar_empresas(self, empresas):
        """Muestra empresas en la tabla"""
        existing_count = len(self.tree.get_children())
        for idx, empresa in enumerate(empresas, start=existing_count):
            empresa_id = empresa.get("id")
            empresa_id_str = str(empresa_id) if empresa_id is not None else ""
            if empresa_id_str:
                self._empresa_por_id[empresa_id_str] = empresa
            row_tag = "oddrow" if idx % 2 else "evenrow"
            tags = (empresa_id_str, row_tag) if empresa_id_str else (row_tag,)
            self.tree.insert("", tk.END, values=(
                empresa.get("nombre_empresa", ""),
                empresa.get("nit_empresa", ""),
                empresa.get("ciudad_empresa", ""),
                empresa.get("estado", ""),
                empresa.get("zona_empresa", ""),
                empresa.get("profesional_asignado", ""),
                empresa.get("asesor", ""),
                empresa.get("contacto_empresa", ""),
                empresa.get("telefono_empresa", ""),
                empresa.get("sede_empresa", ""),
            ), tags=tags)

# ============================================
# THEME HELPERS
# ============================================

_BUTTON_STYLES = {
    "primary": (COLOR_TEAL, "white", COLOR_TEAL_DARK),
    "secondary": (COLOR_PURPLE, "white", COLOR_PURPLE_DARK),
    "danger": (COLOR_DANGER, "white", COLOR_DANGER_DARK),
    "success": (COLOR_SUCCESS, "white", COLOR_SUCCESS_DARK),
    "info": (COLOR_INFO, "white", COLOR_INFO_DARK),
    "warning": (COLOR_WARNING, "white", COLOR_WARNING_DARK),
    "neutral": (COLOR_NEUTRAL, "white", COLOR_NEUTRAL_DARK),
    "outline": (COLOR_WHITE, COLOR_TEXT, COLOR_LIGHT_BG),
}


def _make_button(parent, text, command, style="primary", font=None, **kwargs):
    bg, fg, hover_bg = _BUTTON_STYLES.get(style, _BUTTON_STYLES["primary"])
    btn_font = font if font else FONT_H3
    btn = tk.Button(
        parent,
        text=text,
        command=command,
        font=btn_font,
        bg=bg,
        fg=fg,
        activebackground=hover_bg,
        activeforeground=fg,
        cursor="hand2",
        relief="flat",
        bd=0,
        padx=BTN_PADX,
        pady=BTN_PADY,
        **kwargs,
    )
    if style == "outline":
        btn.config(relief="solid", bd=1, highlightthickness=0)

    def on_enter(_e):
        btn.config(bg=hover_bg)

    def on_leave(_e):
        btn.config(bg=bg)

    btn.bind("<Enter>", on_enter)
    btn.bind("<Leave>", on_leave)
    return btn


def _configure_global_styles(root):
    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except tk.TclError:
        pass

    # Treeview
    style.configure(
        "Treeview",
        rowheight=28,
        font=FONT_BODY,
        background=COLOR_WHITE,
        fieldbackground=COLOR_WHITE,
        borderwidth=0,
    )
    style.configure(
        "Treeview.Heading",
        background=COLOR_PURPLE,
        foreground="white",
        font=FONT_BODY_BOLD,
        borderwidth=0,
        relief="flat",
    )
    style.map(
        "Treeview.Heading",
        background=[("active", COLOR_PURPLE_DARK)],
    )
    style.map(
        "Treeview",
        background=[("selected", "#D4C5E2")],
        foreground=[("selected", COLOR_TEXT)],
    )

    # Progressbar
    style.configure(
        "Reca.Horizontal.TProgressbar",
        background=COLOR_TEAL,
        troughcolor="#EDE7F3",
        bordercolor="#EDE7F3",
        lightcolor=COLOR_TEAL,
        darkcolor=COLOR_TEAL,
    )

    # Combobox
    style.configure("TCombobox", font=FONT_BODY)

    # Scrollbar
    style.configure(
        "TScrollbar",
        width=10,
        troughcolor=COLOR_LIGHT_BG,
        background=COLOR_BORDER,
        borderwidth=0,
    )
    style.map(
        "TScrollbar",
        background=[("active", COLOR_PURPLE), ("!active", COLOR_BORDER)],
    )

    # Entry
    style.configure("TEntry", font=FONT_BODY)

    # Notebook tabs
    style.configure("TNotebook.Tab", font=FONT_BODY, padding=(SP_MD, SP_SM))


# ============================================
# MAIN
# ============================================

if __name__ == "__main__":
    LOG.info("App start version %s", APP_VERSION)
    root = tk.Tk()
    _configure_global_styles(root)
    root.withdraw()
    splash = SplashScreen(root)
    splash.set_status("Preparando...", 5)

    def start_app():
        splash.set_status("Iniciando...", 10)
        root.update_idletasks()
        root.update()

        if not _ensure_credentials():
            messagebox.showerror("Error", "Credenciales no configuradas (.env)")
            splash.close()
            root.destroy()
            sys.exit(1)

        def on_ready():
            splash.close()
            root.deiconify()

        splash.set_status("Iniciando aplicacion...", 20)
        AppMenu(root)
        on_ready()

    root.after(50, start_app)
    root.mainloop()
