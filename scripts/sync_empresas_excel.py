import argparse
import json
import os
import re
import unicodedata
from collections import defaultdict
from datetime import datetime

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client


def to_text(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text if text else None


def strip_invisible_chars(text):
    if text is None:
        return None
    return "".join(
        ch for ch in text if unicodedata.category(ch) not in ("Cf", "Cc") or ch in ("\n", "\t", "\r")
    )


def normalize_spaces(text):
    if text is None:
        return None
    text = strip_invisible_chars(text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_nit(value):
    text = to_text(value)
    if not text:
        return ""
    text = strip_invisible_chars(text)
    text = text.replace(" ", "")
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    return text.lower()


def normalize_name(value):
    text = normalize_spaces(to_text(value))
    return text.lower() if text else ""


def map_excel_row(row):
    return {
        "nombre_empresa": normalize_spaces(to_text(row[0] if len(row) > 0 else None)),
        "nit_empresa": normalize_spaces(to_text(row[1] if len(row) > 1 else None)),
        "direccion_empresa": normalize_spaces(to_text(row[2] if len(row) > 2 else None)),
        "ciudad_empresa": normalize_spaces(to_text(row[3] if len(row) > 3 else None)),
        "correo_1": normalize_spaces(to_text(row[4] if len(row) > 4 else None)),
        "contacto_empresa": normalize_spaces(to_text(row[5] if len(row) > 5 else None)),
        "cargo": normalize_spaces(
            to_text(
                row[10]
                if len(row) > 10 and row[10] not in (None, "")
                else (row[6] if len(row) > 6 else None)
            )
        ),
        "sede_empresa": normalize_spaces(to_text(row[7] if len(row) > 7 else None)),
        "telefono_empresa": normalize_spaces(to_text(row[8] if len(row) > 8 else None)),
        "responsable_visita": normalize_spaces(to_text(row[9] if len(row) > 9 else None)),
        "asesor": normalize_spaces(to_text(row[11] if len(row) > 11 else None)),
        "correo_asesor": normalize_spaces(to_text(row[12] if len(row) > 12 else None)),
        "zona_empresa": normalize_spaces(to_text(row[13] if len(row) > 13 else None)),
        "caja_compensacion": normalize_spaces(to_text(row[14] if len(row) > 14 else None)),
        "profesional_asignado": normalize_spaces(to_text(row[15] if len(row) > 15 else None)),
        "correo_profesional": normalize_spaces(to_text(row[16] if len(row) > 16 else None)),
        "estado": normalize_spaces(to_text(row[17] if len(row) > 17 else None)),
        "observaciones": normalize_spaces(to_text(row[18] if len(row) > 18 else None)),
    }


def load_sheet(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        rec = map_excel_row(row)
        if any(v for v in rec.values()):
            records.append(rec)
    return ws.title, records


def fetch_db_rows(client):
    rows = []
    offset = 0
    batch = 1000
    while True:
        data = (
            client.table("empresas")
            .select("id,nit_empresa,nombre_empresa")
            .range(offset, offset + batch - 1)
            .execute()
            .data
            or []
        )
        rows.extend(data)
        if len(data) < batch:
            break
        offset += batch
    return rows


def build_plan(sheet_rows, db_rows):
    sheet_by_pair = {}
    for rec in sheet_rows:
        key = (normalize_nit(rec.get("nit_empresa")), normalize_name(rec.get("nombre_empresa")))
        if not key[0] and not key[1]:
            continue
        sheet_by_pair[key] = rec

    db_by_pair = defaultdict(list)
    for row in db_rows:
        key = (normalize_nit(row.get("nit_empresa")), normalize_name(row.get("nombre_empresa")))
        db_by_pair[key].append(row)

    update_ops = []
    insert_ops = []
    for key, rec in sheet_by_pair.items():
        matches = db_by_pair.get(key)
        if matches:
            for match in matches:
                update_ops.append((match["id"], rec))
        else:
            insert_ops.append(rec)

    return sheet_by_pair, db_by_pair, update_ops, insert_ops


def backup_current_table(client, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    all_rows = []
    offset = 0
    batch = 1000
    while True:
        data = client.table("empresas").select("*").range(offset, offset + batch - 1).execute().data or []
        all_rows.extend(data)
        if len(data) < batch:
            break
        offset += batch

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(output_dir, f"empresas_backup_{ts}.json")
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(all_rows, fh, ensure_ascii=False, indent=2)
    return path, len(all_rows)


def apply_changes(client, update_ops, insert_ops):
    updated = 0
    inserted = 0

    for row_id, payload in update_ops:
        client.table("empresas").update(payload).eq("id", row_id).execute()
        updated += 1

    chunk_size = 200
    for i in range(0, len(insert_ops), chunk_size):
        chunk = insert_ops[i : i + chunk_size]
        client.table("empresas").insert(chunk).execute()
        inserted += len(chunk)

    return updated, inserted


def main():
    parser = argparse.ArgumentParser(description="Sync Excel companies into Supabase empresas table.")
    parser.add_argument("--excel", required=True, help="Path to XLSX file")
    parser.add_argument("--apply", action="store_true", help="Apply updates/inserts")
    parser.add_argument("--backup-dir", default="backups", help="Backup folder")
    args = parser.parse_args()

    load_dotenv(".env")
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_KEY")
    if not url or not key:
        raise RuntimeError("Missing SUPABASE_URL/SUPABASE_KEY")

    client = create_client(url, key)

    sheet_name, sheet_rows = load_sheet(args.excel)
    db_rows = fetch_db_rows(client)
    sheet_by_pair, db_by_pair, update_ops, insert_ops = build_plan(sheet_rows, db_rows)

    duplicate_db_pairs = sum(1 for key in sheet_by_pair if len(db_by_pair.get(key, [])) > 1)
    same_nit_multi_name = defaultdict(set)
    for nit, name in sheet_by_pair:
        if nit:
            same_nit_multi_name[nit].add(name)
    multi_name_nits = sum(1 for names in same_nit_multi_name.values() if len(names) > 1)

    print("sheet_name=", sheet_name)
    print("sheet_rows=", len(sheet_rows))
    print("sheet_unique_nit_name=", len(sheet_by_pair))
    print("db_rows=", len(db_rows))
    print("update_ops=", len(update_ops))
    print("insert_ops=", len(insert_ops))
    print("duplicate_db_pairs_for_sheet=", duplicate_db_pairs)
    print("nits_with_multiple_names_in_sheet=", multi_name_nits)

    if not args.apply:
        print("mode=dry_run")
        return

    backup_path, backup_rows = backup_current_table(client, args.backup_dir)
    print("backup_path=", backup_path)
    print("backup_rows=", backup_rows)

    updated, inserted = apply_changes(client, update_ops, insert_ops)
    print("applied_updates=", updated)
    print("applied_inserts=", inserted)


if __name__ == "__main__":
    main()
